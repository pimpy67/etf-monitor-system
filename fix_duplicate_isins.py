"""
fix_duplicate_isins.py - Corregge i 15 ISIN sbagliati nel mapping
=================================================================
Aggiorna isin_mapping.json e il file Excel con gli ISIN corretti.
Eseguire una sola volta.
"""

import json
from openpyxl import load_workbook
from openpyxl.styles import Font

MAPPING_PATH = 'data/isin_mapping.json'
EXCEL_PATH = 'etf_monitoraggio.xlsx'

# Correzioni: excel_index -> ISIN corretto + nome JustETF corretto
CORRECTIONS = {
    # GROUP 1 - EUPA.MI (erano tutti Ossiam CAPE LU2555926372)
    48: {
        'isin': 'LU1681042609',
        'justetf_name': 'Amundi MSCI Europe ESG Broad Transition UCITS ETF - EUR (C)',
    },
    53: {
        'isin': 'LU1940199984',
        'justetf_name': 'Amundi MSCI Europe ESG Leaders UCITS ETF EUR Hedged Acc',
    },
    54: {
        'isin': 'LU1940199711',
        'justetf_name': 'Amundi MSCI Europe ESG Selection UCITS ETF Acc',
    },
    70: {
        'isin': 'LU1861137484',
        'justetf_name': 'Amundi MSCI Europe SRI Climate Paris Aligned UCITS ETF DR (C)',
    },
    73: {
        'isin': 'LU2059756598',
        'justetf_name': 'Amundi MSCI Europe SRI Climate Paris Aligned UCITS ETF DR (D)',
    },

    # GROUP 2 - USCP.MI (erano Ossiam CAPE US LU1079841273)
    162: {
        'isin': 'IE0000U24AJ9',
        'justetf_name': 'Amundi MSCI USA SRI Climate Paris Aligned UCITS ETF Acc EUR Hedged',
    },
    163: {
        'isin': 'IE000R85HL30',
        'justetf_name': 'Amundi MSCI USA SRI Climate Paris Aligned UCITS ETF Acc',
    },

    # GROUP 3 - ECRP/CORP (erano tutti LU1525418643, tranne index 107 che era corretto)
    87: {
        'isin': 'LU1681039647',
        'justetf_name': 'Amundi Index Euro Corporate SRI UCITS ETF 2 DR EUR (C)',
    },
    90: {
        'isin': 'LU1437018168',
        'justetf_name': 'Amundi Index Euro Corporate SRI UCITS ETF DR (C)',
    },
    # 107: GIA' CORRETTO (LU1525418643)
    130: {
        'isin': 'LU2037748774',
        'justetf_name': 'Amundi Index Euro Corporate SRI 0-3 Y UCITS ETF DR (C)',
    },
    144: {
        'isin': 'LU1525418726',
        'justetf_name': 'Amundi Global Corporate SRI 1-5Y UCITS ETF DR (C)',
    },

    # GROUP 4 - Singoli match sbagliati
    11: {
        'isin': 'LU1834988609',
        'justetf_name': 'Amundi STOXX Europe 600 Telecommunications UCITS ETF Acc',
    },
    63: {
        'isin': 'LU1602144575',
        'justetf_name': 'Amundi MSCI EMU ESG Selection UCITS ETF DR - EUR (C)',
    },
    167: {
        'isin': 'IE0008TKP6O7',
        'justetf_name': 'Amundi MSCI USA ESG Leaders Extra UCITS ETF DR - USD (D)',
    },
    76: {
        'isin': 'LU1650491282',
        'justetf_name': 'Amundi Euro Government Inflation-Linked Bond UCITS ETF Acc',
    },
}


def fix_mapping():
    """Aggiorna il file isin_mapping.json"""
    with open(MAPPING_PATH, 'r') as f:
        mapping = json.load(f)

    fixed = 0
    for entry in mapping:
        idx = entry.get('excel_index')
        if idx in CORRECTIONS:
            old_isin = entry['isin']
            correction = CORRECTIONS[idx]
            entry['isin'] = correction['isin']
            entry['justetf_name'] = correction['justetf_name']
            entry['confidence'] = 1.0
            entry['method'] = 'manual_fix'
            print(f"  [{idx:3d}] {entry['excel_ticker']:10s} {old_isin} -> {correction['isin']}")
            print(f"        {entry['excel_name'][:55]}")
            fixed += 1

    with open(MAPPING_PATH, 'w') as f:
        json.dump(mapping, f, indent=2, ensure_ascii=False)

    print(f"\nAggiornati {fixed} ISIN nel mapping JSON")
    return mapping


def fix_excel(mapping):
    """Aggiorna la colonna ISIN nel file Excel"""
    wb = load_workbook(EXCEL_PATH)
    ws = wb['ETF']

    # Trova colonna ISIN
    isin_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == 'ISIN':
            isin_col = col
            break

    if not isin_col:
        print("Colonna ISIN non trovata nell'Excel!")
        return

    print(f"\nColonna ISIN trovata: colonna {isin_col}")

    fixed = 0
    for entry in mapping:
        idx = entry.get('excel_index')
        if idx in CORRECTIONS:
            row = idx + 2  # +1 header, +1 zero-indexed
            new_isin = CORRECTIONS[idx]['isin']
            ws.cell(row=row, column=isin_col, value=new_isin)
            ws.cell(row=row, column=isin_col).font = Font(color="006600")
            fixed += 1

    wb.save(EXCEL_PATH)
    print(f"Aggiornati {fixed} ISIN nel file Excel")


def verify(mapping):
    """Verifica che non ci siano piu' ISIN duplicati problematici"""
    isins = [m['isin'] for m in mapping if m.get('isin')]
    from collections import Counter
    counts = Counter(isins)
    duplicates = {isin: count for isin, count in counts.items() if count > 1}

    if duplicates:
        print(f"\nISIN ancora duplicati ({len(duplicates)} gruppi):")
        for isin, count in sorted(duplicates.items(), key=lambda x: -x[1]):
            entries = [m for m in mapping if m.get('isin') == isin]
            names = [e['excel_name'][:40] for e in entries]
            print(f"  {isin} ({count}x): {', '.join(names)}")
    else:
        print("\nNessun ISIN duplicato! Tutti unici.")


def main():
    print("=" * 60)
    print("FIX ISIN DUPLICATI")
    print("=" * 60)
    print(f"\nCorrezioni da applicare: {len(CORRECTIONS)}")

    # 1. Fix mapping JSON
    print("\n--- Aggiornamento isin_mapping.json ---")
    mapping = fix_mapping()

    # 2. Fix Excel
    print("\n--- Aggiornamento Excel ---")
    fix_excel(mapping)

    # 3. Verifica
    print("\n--- Verifica ---")
    verify(mapping)

    print("\n" + "=" * 60)
    print("Correzione completata!")
    print("=" * 60)


if __name__ == '__main__':
    main()
