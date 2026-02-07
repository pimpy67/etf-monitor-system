"""
migrate_to_isin.py - Script di migrazione per aggiungere ISIN all'Excel
=========================================================================
Esegue:
1. Risoluzione ISIN per tutti gli ETF (se non gia' fatto)
2. Aggiorna le intestazioni Excel (ADX -> MACD, Vol Ratio -> BB Width)
3. Aggiorna il file Excel con colonna ISIN

Eseguire una sola volta prima del deploy.
"""

import json
import os
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


EXCEL_PATH = 'etf_monitoraggio.xlsx'
MAPPING_PATH = 'data/isin_mapping.json'


def ensure_isin_mapping():
    """Genera il mapping ISIN se non esiste"""
    if os.path.exists(MAPPING_PATH):
        print(f"Mapping ISIN gia' presente: {MAPPING_PATH}")
        with open(MAPPING_PATH, 'r') as f:
            return json.load(f)

    print("Mapping ISIN non trovato. Eseguo risoluzione...")
    from isin_resolver import ISINResolver
    resolver = ISINResolver()
    results = resolver.resolve_all(EXCEL_PATH)
    resolver.save_mapping(results, MAPPING_PATH)
    return results


def update_excel_headers(wb):
    """Aggiorna le intestazioni Excel: ADX -> MACD Hist, Vol Ratio -> BB Width"""
    ws = wb['ETF']

    # Riga 1 = intestazioni
    header_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            header_map[val] = col

    # Rinomina ADX -> MACD Hist
    if 'ADX' in header_map:
        col = header_map['ADX']
        ws.cell(row=1, column=col, value='MACD Hist')
        print(f"  Intestazione colonna {col}: ADX -> MACD Hist")

    # Rinomina Vol Ratio -> BB Width
    if 'Vol Ratio' in header_map:
        col = header_map['Vol Ratio']
        ws.cell(row=1, column=col, value='BB Width')
        print(f"  Intestazione colonna {col}: Vol Ratio -> BB Width")

    return wb


def add_isin_to_excel(mapping):
    """Aggiunge/aggiorna la colonna ISIN nel file Excel"""
    print(f"\nAggiornamento file Excel: {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH)
    ws = wb['ETF']

    # Controlla se la colonna ISIN esiste gia'
    header_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            header_map[val] = col

    # Aggiorna intestazioni (ADX -> MACD, Vol Ratio -> BB Width)
    wb = update_excel_headers(wb)

    if 'ISIN' in header_map:
        isin_col = header_map['ISIN']
        print(f"  Colonna ISIN gia' presente (colonna {isin_col}), aggiorno valori...")
    else:
        # Aggiungi ISIN come ultima colonna
        isin_col = ws.max_column + 1
        ws.cell(row=1, column=isin_col, value='ISIN')
        ws.cell(row=1, column=isin_col).font = Font(bold=True)
        print(f"  Aggiunta colonna ISIN (colonna {isin_col})")

    # Popola ISIN
    updated = 0
    for entry in mapping:
        idx = entry.get('excel_index', -1)
        isin = entry.get('isin', '')
        confidence = entry.get('confidence', 0)

        if idx >= 0 and isin:
            row = idx + 2  # +1 per header, +1 per 0-indexed
            ws.cell(row=row, column=isin_col, value=isin)

            # Colora in base alla confidenza
            if confidence >= 0.7:
                ws.cell(row=row, column=isin_col).font = Font(color="006600")
            else:
                ws.cell(row=row, column=isin_col).font = Font(color="CC0000", bold=True)
            updated += 1

    wb.save(EXCEL_PATH)
    print(f"  Aggiornati {updated} ISIN nel file Excel")


def print_summary(mapping):
    """Stampa riepilogo della migrazione"""
    total = len(mapping)
    high_conf = sum(1 for m in mapping if m.get('confidence', 0) >= 0.7)
    low_conf = sum(1 for m in mapping if 0 < m.get('confidence', 0) < 0.7)
    not_found = sum(1 for m in mapping if m.get('confidence', 0) == 0)

    # Conta ISIN unici
    isins = [m['isin'] for m in mapping if m.get('isin')]
    unique_isins = len(set(isins))
    duplicate_isins = total - unique_isins

    print(f"\n{'='*60}")
    print(f"RIEPILOGO MIGRAZIONE")
    print(f"{'='*60}")
    print(f"  Totale ETF:            {total}")
    print(f"  ISIN trovati (alta):   {high_conf}")
    print(f"  ISIN trovati (bassa):  {low_conf}")
    print(f"  ISIN non trovati:      {not_found}")
    print(f"  ISIN unici:            {unique_isins}")
    print(f"  ISIN duplicati:        {duplicate_isins}")

    if duplicate_isins > 0:
        print(f"\n  NOTA: Ci sono {duplicate_isins} righe con ISIN duplicati.")
        print(f"  Questi potrebbero essere ETF diversi mappati allo stesso fondo,")
        print(f"  oppure match sbagliati da correggere manualmente in:")
        print(f"  {MAPPING_PATH}")

    if low_conf > 0:
        print(f"\n  ATTENZIONE: {low_conf} ETF con bassa confidenza:")
        for m in mapping:
            if 0 < m.get('confidence', 0) < 0.7:
                print(f"    {m['excel_ticker']:15s} -> {m['isin']:14s} (conf={m['confidence']:.2f})")
                print(f"      Excel: {m['excel_name'][:50]}")
                print(f"      Match: {m['justetf_name'][:50]}")

    print(f"\n{'='*60}")
    print(f"Migrazione completata!")
    print(f"Prossimi passi:")
    print(f"  1. Controlla i match a bassa confidenza in {MAPPING_PATH}")
    print(f"  2. Correggi manualmente se necessario")
    print(f"  3. Fai deploy su Railway")
    print(f"{'='*60}")


def main():
    """Esegui migrazione completa"""
    print("=" * 60)
    print("MIGRAZIONE ETF MONITOR -> ISIN + JustETF")
    print("=" * 60)

    # 1. Genera/carica mapping ISIN
    mapping = ensure_isin_mapping()

    # 2. Aggiorna Excel
    add_isin_to_excel(mapping)

    # 3. Riepilogo
    print_summary(mapping)


if __name__ == "__main__":
    main()
