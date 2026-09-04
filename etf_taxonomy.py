"""
Tassonomia ETF — classificazione multi-dimensionale (Asset Class / Geografia / Settore).

Serve SOLO alla vista "Momentum Settori" della dashboard. NON tocca `detect_family()`
né la colonna `Categoria` dell'Excel (che resta la fonte per l'engine L0/L1).

Uso come libreria:
    from etf_taxonomy import classify
    ac, geo, sett = classify(nome_etf, categoria_excel)

Uso come CLI:
    python etf_taxonomy.py --preview            # stampa CSV su stdout
    python etf_taxonomy.py --write              # scrive/aggiorna le colonne
                                                #   'Asset Class' / 'Geografia' / 'Settore'
                                                #   nel foglio 'ETF' di etf_monitoraggio.xlsx

Le tre colonne nell'Excel sono un OVERRIDE editabile a mano: se una cella è compilata
il monitor la usa così com'è, se è vuota il monitor richiama classify() al volo
(vedi monitor.py::analyze_etf).
"""
import re
import sys

EXCEL_PATH_DEFAULT = 'etf_monitoraggio.xlsx'
NEW_COLS = ('Asset Class', 'Geografia', 'Settore')

_EQ_IDX = (r'stoxx|msci|s&p ?\d|s&p ?5|nasdaq|ftse ?mib|ftse ?100|ftse ?italia|'
           r'ftse all-?world|\bdax\b|\bcac\b|topix|nikkei|russell|dow jones|dj |titans|'
           r'epra|nareit|smi\b|catholic|equal[- ]weight equ')


def _has(s, *pats):
    return any(re.search(p, s) for p in pats)


def asset_class(nome: str, categoria: str) -> str:
    s = (nome + ' ' + categoria).lower()
    if _has(s, r'crypto|bitcoin|ethereum|\bsolana\b|digital asset'):
        return 'Crypto'
    if _has(s, r'\bbond\b|oblig|govern|treasur|\bgilt|\bbund\b|\bbtp\b|aggregate|corporate|'
               r'corp\.|high yield|\bhy\b|fixed income|inflation|floating rate|\bcash\b|'
               r'overnight|money market|monetar|liquidit'):
        return 'Bond'
    if _has(s, r'\bcrypto\b'):
        return 'Crypto'
    if _has(s, _EQ_IDX):
        return 'Equity'
    if _has(s, r'physical (gold|silver|platinum|palladium)|\betc\b|bloomberg comm|'
               r'equal-weight comm|industrial metals|broad commod|\buranium\b|crude oil|'
               r'natural gas|agricultur'):
        return 'Commodity'
    if _has(s, r'multi-asset|multi asset|allocation|balanced|bilanciat|target risk'):
        return 'Multi-asset'
    return 'Equity'


def geografia(nome: str, categoria: str, ac: str) -> str:
    n = nome.lower()
    n = re.sub(r'eur h(dg|edged)|usd hedged|daily hedged|screened|esg|sri|'
              r'climate paris aligned|broad transition|selection|swap( ii)?|ucits etf.*$',
              ' ', n)
    c = categoria.lower()
    if _has(n, r'\bitaly\b|\bitalia\b|ftse ?mib|\bmib\b|\bbtp\b'):
        return 'Italia'
    if _has(n, r'\bjapan\b|giappone|topix|nikkei') and not _has(n, r'ex[- ]?japan'):
        return 'Giappone'
    if _has(n, r'ex[- ]?japan|asia pacific|asia ex|\basia\b|\bchina\b|\bcina\b|\bindia\b|'
               r'\bkorea\b|taiwan|indonesia|\bbrazil\b|brasil|latin america|emerging|'
               r'frontier|\bem \b|pan africa|middle east|greece|grecia|eastern europe'):
        return 'EM'
    if _has(n, r'united kingdom|\buk\b|ftse ?100|regno unito'):
        return 'Regno Unito'
    if _has(n, r'\bcanada\b'):
        return 'Canada'
    if _has(n, r'australia|s&p/asx'):
        return 'Asia Sviluppata'
    if _has(n, r'all[- ]?world|\bacwi\b|all country world|\bworld\b|global|'
               r'developed markets|global titans|dj global'):
        return 'Globale'
    if _has(n, r's&p ?500|s&p ?5|msci usa|nasdaq|russell ?2000|dow jones|\bus \b|\busa\b|'
               r'u\.s\.|united states'):
        return 'USA'
    if _has(n, r'euro stoxx|eurozone|\bemu\b|europe|europa|stoxx europe|\bdax\b|\bcac\b|'
               r'nordic|\bsmi\b|swiss|switzerland|\beuro \b|\beur \b'):
        return 'Europa'
    if ac == 'Bond':
        if _has(n, r'\$ treas|usd treas|us treas|usd corp|usd hy|usd high yield|usd float'):
            return 'USA'
        if _has(n, r'global|aggregate'):
            return 'Globale'
        if _has(n, r'\beuro\b|\beur\b|italy|btp'):
            return 'Europa'
    if ac == 'Commodity':
        return 'Globale'
    if _has(n, r'\$ trea|us trea'):
        return 'USA'
    if 'stati uniti' in c:
        return 'USA'
    if 'europa occidentale' in c or 'azionari europa' in c:
        return 'Europa'
    if 'asia - emergente' in c or 'america latina' in c or "europa dell" in c or 'africa' in c:
        return 'EM'
    if 'giappone' in c:
        return 'Giappone'
    if 'asia - paesi sviluppati' in c or 'australasia' in c:
        return 'Asia Sviluppata'
    if 'globale' in c or 'all-world' in c or 'azionari sviluppati' in c:
        return 'Globale'
    if 'regno unito' in c:
        return 'Regno Unito'
    if 'canada' in c:
        return 'Canada'
    if 'azionari italia' in c:
        return 'Italia'
    if ac in ('Equity', 'Commodity', 'Crypto'):
        return 'Globale'
    return 'Europa' if ac == 'Bond' else 'n/d'


def settore(nome: str, categoria: str, ac: str) -> str:
    s = (nome + ' ' + categoria).lower()
    ln = nome.lower()
    if ac in ('Bond', 'Monetario', 'Multi-asset'):
        return '—'
    if ac == 'Crypto':
        return 'Crypto'
    if _has(s, r'\d ?x (long|short).*(nvidia|tesla|apple|amazon|meta|microsoft|nvda)|3x (long|short)'):
        return 'Leva Titolo Singolo'
    if _has(ln, r'timber|forestry'):
        return 'Materiali'
    if _has(ln, r'vix futures') or (_has(ln, r'volatilit') and not _has(ln, r'min(imum)? vol')):
        return 'Volatilità'
    if _has(s, r'semicondutt|semiconductor|memory chip|\bchips?\b'):
        return 'Semiconduttori'
    if _has(s, r'artificial intelligence|\ba\.?i\.?\b|big data|robotic|automation'):
        return 'AI & Robotics'
    if _has(s, r'cyber|\bsecurity\b'):
        return 'Cybersecurity'
    if _has(s, r'clean energy|new energy|renewable|solar|\bwind\b|hydrogen|energy transition|'
               r'batter|bioenerg'):
        return 'Clean Energy & Battery'
    if _has(s, r'technolog|software|digital econom|\binternet\b|e-commerce|fintech|\bcloud\b|'
               r'digitalis|smart mobility|millennials'):
        return 'Technology'
    if _has(s, r'health ?care|\bhealth\b|salute|pharma|biotech|medical|genomic|farmac'):
        return 'Healthcare'
    if _has(s, r'\bbank|insurance|\bfinancial|finanza'):
        return 'Finanza'
    if _has(s, r'real est|\breit|nareit|\bepra\b|immobil|property'):
        return 'Real Estate'
    if _has(s, r'telecom|communication serv'):
        return 'Telecom'
    if _has(s, r'utilit|\bwater\b|\bacqua\b'):
        return 'Utilities & Water'
    if _has(s, r'infrastruct|infrastruttur'):
        return 'Infrastrutture'
    if _has(s, r'industrial metal|metalli industriali|\bcopper\b|\bnickel\b|\bzinc\b|alumini|'
               r'lithium|rame'):
        return 'Metalli Industriali'
    if _has(s, r'defen[cs]e|difesa|aerospace'):
        return 'Difesa'
    if _has(s, r'industrial|industria'):
        return 'Industria'
    if _has(s, r'consumer|consum|luxury|lusso|retail|\bfood\b|beverage|nutrition'):
        return 'Consumi'
    if _has(s, r'\bgold\b|silver|platinum|palladium|precious metal|\boro\b|argento'):
        return 'Oro & Metalli Preziosi'
    if _has(s, r'\benergy\b|energia|\boil\b|\bgas\b|petrol'):
        return 'Energia'
    if _has(s, r'basic res|basic material|\bmaterials\b|materiali|timber|forestry|\bmining\b|chemical'):
        return 'Materiali'
    if ac == 'Commodity':
        return 'Commodity Broad'
    return 'Broad'


def classify(nome: str, categoria: str) -> dict:
    """Ritorna {'asset_class', 'geografia', 'settore'} da nome ETF + categoria Excel."""
    nome = str(nome or '')
    categoria = str(categoria or '')
    ac = asset_class(nome, categoria)
    return {
        'asset_class': ac,
        'geografia': geografia(nome, categoria, ac),
        'settore': settore(nome, categoria, ac),
    }


# ─────────────────────────────── CLI ────────────────────────────────
def _iter_rows(ws):
    header = [c.value for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}
    for r in ws.iter_rows(min_row=2):
        yield r, idx


def _main(argv):
    import openpyxl
    import csv
    path = EXCEL_PATH_DEFAULT
    write = '--write' in argv
    for a in argv:
        if a.startswith('--path='):
            path = a.split('=', 1)[1]

    wb = openpyxl.load_workbook(path)
    ws = wb['ETF']
    header = [c.value for c in ws[1]]
    # aggiungi le colonne mancanti in coda
    for col in NEW_COLS:
        if col not in header:
            ws.cell(row=1, column=len(header) + 1, value=col)
            header.append(col)
    idx = {name: i + 1 for i, name in enumerate(header)}  # 1-based

    out = csv.writer(sys.stdout)
    out.writerow(['Ticker', 'Nome', 'Categoria', *NEW_COLS])
    n = 0
    for row in ws.iter_rows(min_row=2):
        nome = row[idx['Nome ETF'] - 1].value
        if not nome:
            continue
        cat = row[idx['Categoria'] - 1].value or ''
        res = classify(nome, cat)
        vals = (res['asset_class'], res['geografia'], res['settore'])
        out.writerow([row[idx['Ticker'] - 1].value, nome, cat, *vals])
        if write:
            for col, v in zip(NEW_COLS, vals):
                ws.cell(row=row[0].row, column=idx[col], value=v)
        n += 1

    if write:
        wb.save(path)
        print(f'\n[scritte {n} righe → {path}]', file=sys.stderr)
    else:
        print(f'\n[{n} righe — anteprima, nessuna scrittura. usa --write]', file=sys.stderr)


if __name__ == '__main__':
    _main(sys.argv[1:])
