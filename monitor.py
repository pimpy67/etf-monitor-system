"""
monitor.py - Monitoraggio ETF
==============================
Orchestrazione del sistema:
1. Legge file Excel con lista ETF (ticker Yahoo Finance + ISIN)
2. Recupera dati OHLCV via Yahoo Finance
3. Calcola indicatori tecnici (EMA20, SMA50, SMA200, ADX, RSI)
4. Gestisce livelli L0/L1/L2/L3
5. Invia alert giornalieri
6. Genera dashboard JSON
"""

import os
import sys
from datetime import datetime, timedelta, date as date_type
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import json
import time
import numpy as np
from decimal import Decimal
from pathlib import Path
import traceback

from data_fetcher import ETFDataFetcher
from technical_analysis import ETFTechnicalAnalyzer
from alerts import AlertSystem
from database import PriceDatabase
from pdf_generator import generate_parameters_pdf
from pdf_generator_complete import generate_complete_pdf
from order_pricing import compute_order_prices


monitor_log = []


def add_log(message: str):
    entry = f"[{datetime.now().strftime('%H:%M:%S')}] {message}"
    monitor_log.append(entry)
    if len(monitor_log) > 200:
        monitor_log.pop(0)
    print(entry)


class ETFMonitor:
    """Sistema principale di monitoraggio ETF"""

    def __init__(self, excel_path: str = 'etf_monitoraggio.xlsx'):
        self.excel_path   = excel_path
        self.data_fetcher = ETFDataFetcher(rate_limit=1.0)
        self.alert_system = AlertSystem()
        self.db           = PriceDatabase()
        # PRIORITÀ 3 — L2 Readiness anti-flickering state
        self.l2_score_state = {}  # {isin: {'raw': score, 'smoothed': ema_score, 'raw_prev': prev_score}}
        self._load_l2_state()

    def _load_l2_state(self):
        """Carica stato persistente L2 smoothing da file."""
        try:
            import json
            if os.path.exists('data/l2_score_state.json'):
                with open('data/l2_score_state.json', 'r') as f:
                    self.l2_score_state = json.load(f)
        except Exception:
            self.l2_score_state = {}

    def _save_l2_state(self):
        """Salva stato persistente L2 smoothing su file."""
        try:
            import json
            os.makedirs('data', exist_ok=True)
            with open('data/l2_score_state.json', 'w') as f:
                json.dump(self.l2_score_state, f)
        except Exception:
            pass

    def _apply_l2_smoothing(self, isin: str, raw_score: float) -> dict:
        """
        PRIORITÀ 3 — Applica smoothing EMA + isteresi + jump override a L2 score.

        Returns: {
            'raw': raw_score,
            'smoothed': ema_score,
            'enters_watchlist': bool,  # Isteresi: raw >= 70 ma smoothed ancora < 70
            'exits_watchlist': bool,   # Isteresi: raw < 60 ma smoothed ancora >= 60
            'jump_triggered': bool,    # Override salto: delta > 25
        }
        """
        state = self.l2_score_state.get(isin, {'raw': 0, 'smoothed': 0, 'raw_prev': 0})
        raw_prev = state.get('raw_prev', 0)
        smoothed_prev = state.get('smoothed', 0)

        # ── Smoothing EMA (periodo 3gg) ────────────────────────────────
        ema_period = 3
        alpha = 2 / (ema_period + 1)
        if smoothed_prev == 0:
            smoothed = raw_score  # First value
        else:
            smoothed = smoothed_prev + alpha * (raw_score - smoothed_prev)

        # ── Jump override (delta > 25) ────────────────────────────────
        l2_jump_threshold = 25
        delta = raw_score - raw_prev
        jump_triggered = False
        if abs(delta) > l2_jump_threshold and raw_score >= 70:
            smoothed = raw_score  # Hard-reset smoothed al valore grezzo
            jump_triggered = True

        # ── Isteresi (70 enter, 60 exit) ───────────────────────────────
        l2_enter = 70
        l2_exit = 60
        enters = raw_score >= l2_enter and smoothed_prev < l2_enter
        exits = raw_score < l2_exit and smoothed_prev >= l2_exit

        # Salva stato per prossimo ciclo
        self.l2_score_state[isin] = {
            'raw': raw_score,
            'smoothed': smoothed,
            'raw_prev': raw_score
        }

        return {
            'raw': raw_score,
            'smoothed': smoothed,
            'enters_watchlist': enters,
            'exits_watchlist': exits,
            'jump_triggered': jump_triggered,
        }

    def load_etfs(self) -> pd.DataFrame:
        """Carica lista ETF dal file Excel."""
        try:
            df = pd.read_excel(self.excel_path, sheet_name='ETF')
            add_log(f"Caricati {len(df)} ETF dal file Excel")
            return df
        except Exception as e:
            add_log(f"ERRORE caricamento Excel: {e}")
            return pd.DataFrame()

    def get_etf_history(self, ticker: str, isin: str = '') -> pd.DataFrame:
        """
        Recupera storico OHLCV per un ETF.
        Prima prova il DB; se i dati sono datati (> 5gg) rifresca da Yahoo Finance.
        """
        import pandas as _pd

        # 1. Database
        db_df = _pd.DataFrame()
        if isin:
            db_df = self.db.get_ohlc_by_isin(isin, days=260)

        # Usa il DB solo se ha già i dati di oggi (last_date == today)
        if not db_df.empty and len(db_df) >= 55:
            today     = _pd.Timestamp.today().normalize()
            last_date = db_df.index[-1] if hasattr(db_df.index[-1], 'date') else _pd.Timestamp(db_df.index[-1])
            if (today - last_date).days <= 0:
                return db_df  # dati già aggiornati a oggi → no fetch Yahoo Finance

        # 2. Yahoo Finance (OHLCV completo) — DB assente o datato
        df = self.data_fetcher.get_historical_data(ticker, days=260)
        if not df.empty:
            # Yahoo Finance fornisce sempre OHLCV completo: salvalo con save_ohlcv_bulk
            # (save_close_bulk scartava High/Low/Open/Volume anche quando disponibili)
            if isin:
                self.db.save_ohlcv_bulk(isin, df, source='yfinance', isin=isin)
            else:
                self.db.save_ohlcv_bulk(ticker, df, source='yfinance')
            return df

        # 3. DB come fallback anche se datato
        if not db_df.empty:
            return db_df

        return _pd.DataFrame()

    def analyze_etf(self, row: pd.Series) -> dict:
        """Analizza un singolo ETF dalla riga Excel."""
        ticker   = str(row.get('Ticker', '')).strip()
        isin     = str(row.get('ISIN', '')).strip()
        nome     = str(row.get('Nome ETF', ticker))
        categoria = str(row.get('Categoria', ''))
        borsa    = str(row.get('Borsa', ''))
        level    = int(row.get('Livello', 3))

        # Nuovi sistema: usa famiglia da config YAML
        famiglia = ETFTechnicalAnalyzer.detect_family(categoria)
        analyzer = ETFTechnicalAnalyzer(famiglia=famiglia)

        # Mantieni etf_type per backward compatibility
        etf_type = famiglia

        identifier = ticker if ticker else isin
        if not identifier:
            add_log(f"  SKIP: nessun ticker per {nome}")
            return self._empty_result(ticker, isin, nome, categoria, borsa, level,
                                      'Ticker mancante')

        add_log(f"  Analisi {nome[:40]} ({identifier})...")

        hist = self.get_etf_history(ticker, isin)

        if hist.empty or len(hist) < 20:
            add_log(f"    Storico insufficiente: {len(hist)} giorni")
            return self._empty_result(ticker, isin, nome, categoria, borsa, level,
                                      f'Dati insufficienti: {len(hist)} giorni')

        try:
            analysis = analyzer.analyze_etf(hist, current_level=level, ticker=ticker)
        except Exception as e:
            add_log(f"  ERR {identifier}: {type(e).__name__}: {e}")
            import traceback
            add_log(traceback.format_exc())
            return self._empty_result(ticker, isin, nome, categoria, borsa, level,
                                      f'{type(e).__name__}: {e}')

        # STEP 9 — Detecta segnali L0 durante l'analisi (con REGIME FILTER first!)
        l0_entry_signal = {}
        l0_regime_filter = {}
        try:
            if level != 0:  # Se non già in L0
                close_series = hist['Close'] if 'Close' in hist else hist.iloc[:, 0]
                rsi_14 = analysis.get('rsi')
                ema_fast_period = 10  # default; potrebbe venire dalla famiglia
                sma200 = analysis.get('sma200')
                atr_60 = analysis.get('atr', 0) if analysis.get('atr') else None
                volume_20ma = hist['Volume'].rolling(window=20).mean().iloc[-1] if 'Volume' in hist else None

                # 🔴 STEP 9a — REGIME FILTER PRIMA (blocca L0 in regime BULL)
                l0_regime_filter = analyzer.l0_detect_regime_filter(close_series, sma200, atr_60, volume_20ma)
                
                if not l0_regime_filter.get('regime_suitable'):
                    # ❌ Regime NON suitable (BULL market, no bear conditions)
                    add_log(f"    ⛔ L0 BLOCCATO: regime NON suitable ({l0_regime_filter.get('regime_type')}) | days_below_sma200={l0_regime_filter.get('days_below_sma200')}")
                    l0_entry_signal = {
                        'l0_signal': False,
                        'reason': f"REGIME_NOT_SUITABLE: {l0_regime_filter.get('regime_type')}",
                        'regime_days_below_sma200': l0_regime_filter.get('days_below_sma200'),
                    }
                else:
                    # ✅ Regime suitable — ora check le 4 condizioni L0
                    l0_check = analyzer.check_l0_entry(close_series, rsi_14, ema_fast_period)
                    if l0_check.get('l0_signal'):
                        l0_entry_signal = {
                            'l0_signal': True,
                            'dd_from_high': l0_check.get('dd_from_high'),
                            'rsi': l0_check.get('rsi_current'),
                            'recovery': l0_check.get('recovery_pct'),
                            'alert_only': l0_check.get('alert_only', True),
                            'regime_mode': l0_regime_filter.get('regime_type'),  # 'fast_crash' or 'slow_bear'
                            'regime_days_below_sma200': l0_regime_filter.get('days_below_sma200'),
                        }
                    else:
                        l0_entry_signal = {
                            'l0_signal': False,
                            'reason': 'REGIME_OK_BUT_CONDITIONS_NOT_MET',
                            'regime_mode': l0_regime_filter.get('regime_type'),
                            'regime_days_below_sma200': l0_regime_filter.get('days_below_sma200'),
                        }
        except Exception as e:
            add_log(f"    ⚠️  L0 check error: {e}")

        # STEP 10 — Valuta ingresso L1 con logica tiered (Gate + Quality + Size)
        l1_tiered = {}
        l1_accelerated = {}
        try:
            market_data_tiered = {
                'price': analysis.get('current_price'),
                'ema20': analysis.get('ema20'),
                'sma50': analysis.get('sma50'),
                'rsi': analysis.get('rsi'),
                'adx': analysis.get('adx'),
                'macd_h': analysis.get('macd_histogram'),
                'dist_ema20': analysis.get('dist_ema20', 0.0),  # già in percentuale
            }
            l1_tiered = analyzer.check_l1_entry_tiered(market_data_tiered)
            if l1_tiered.get('should_enter'):
                add_log(f"    🟢 L1 TIERED ENTRY: confidence {l1_tiered['confidence']:.0%} | quality {l1_tiered['quality_score']}/4")
        except Exception as e:
            add_log(f"    ⚠️  L1 tiered check error: {e}")

        # STEP 12+ — Valuta ingresso L1 con logica accelerated (Gerarchia 2+2)
        try:
            market_data_accelerated = {
                'close': analysis.get('current_price'),
                'ema20': analysis.get('ema20'),
                'sma50': analysis.get('sma50'),  # Nuovo: usato in parametro X della gerarchia
                'macd_h': analysis.get('macd_histogram'),
                'rsi_14': analysis.get('rsi'),
                'rsi_14_prev': analysis.get('rsi_prev'),
                'adx_14': analysis.get('adx'),
                'adx_14_prev': analysis.get('adx_prev'),
                'days_above_ema20': analysis.get('days_above_ema20', 0),
                'dist_ema20': analysis.get('dist_ema20', 0.0),
            }
            l1_accelerated = analyzer.check_l1_entry_accelerated(market_data_accelerated)
            if l1_accelerated.get('should_enter'):
                add_log(f"    ⚡ L1 GERARCHIA 2+2: gate A∧M ✓ | velocity {l1_accelerated['velocity_count']}/4 ✓ | {l1_accelerated['reason']}")
        except Exception as e:
            add_log(f"    ⚠️  L1 accelerated check error: {e}")

        # STEP 13 — NUOVO: L0 Regime Filter + Doppio Percorso (v4.0)
        l0_regime_filter = {}
        try:
            close_series = hist['Close'] if 'Close' in hist else hist.iloc[:, 0]
            sma200 = analysis.get('sma200')
            atr_60 = analysis.get('atr', 0) if analysis.get('atr') else None
            volume_20ma = hist['Volume'].rolling(window=20).mean().iloc[-1] if 'Volume' in hist else None

            l0_regime_filter = analyzer.l0_detect_regime_filter(close_series, sma200, atr_60, volume_20ma)
            if l0_regime_filter.get('regime_suitable'):
                add_log(f"    📍 L0 REGIME: {l0_regime_filter.get('regime_type')} | "
                       f"days_below_sma200={l0_regime_filter.get('days_below_sma200')} | "
                       f"dd={l0_regime_filter.get('recent_dd_pct')*100:.1f}%")
        except Exception as e:
            add_log(f"    ⚠️  L0 regime filter error: {e}")

        # STEP 14 — NUOVO: L1 Semplificato (7 condizioni tutte obbligatorie, v4.0)
        l1_seven_conditions = {}
        try:
            close_series = hist['Close'] if 'Close' in hist else hist.iloc[:, 0]
            high_series = hist['High'] if 'High' in hist else close_series
            low_series = hist['Low'] if 'Low' in hist else close_series
            macd_prev = analysis.get('macd_histogram_prev', analysis.get('macd_histogram'))
            volume = hist['Volume'].iloc[-1] if 'Volume' in hist else None
            volume_20ma = hist['Volume'].rolling(window=20).mean().iloc[-1] if 'Volume' in hist else None
            atr_14 = analysis.get('atr')

            l1_seven_conditions = analyzer.l1_check_7_conditions(
                close_series, analysis.get('ema20'), analysis.get('sma50'),
                analysis.get('rsi'), analysis.get('adx'),
                analysis.get('macd_histogram'), macd_prev,
                volume, atr_14, high_series, low_series,
                volume_20ma, analysis.get('days_above_ema20', 0)
            )
            if l1_seven_conditions.get('entry_l1'):
                add_log(f"    🔷 L1 7/7 CONDITIONS: ALL TRUE | space={l1_seven_conditions['space_detail'].get('method')}")
        except Exception as e:
            add_log(f"    ⚠️  L1 7/7 check error: {e}")

        # STEP 15 — NUOVO: L2 Readiness Score (pre-screening, PRIORITÀ 3 v4.0)
        l2_readiness_score = 0.0
        l2_smoothed_score = 0.0
        try:
            close_series = hist['Close'] if 'Close' in hist else hist.iloc[:, 0]
            volume = hist['Volume'].iloc[-1] if 'Volume' in hist else None
            volume_20ma = hist['Volume'].rolling(window=20).mean().iloc[-1] if 'Volume' in hist else None
            ema20_series = analysis.get('ema20_series')
            adx_series = analysis.get('adx_series')

            # Calcola raw score con 6 componenti
            l2_readiness_score = analyzer.l2_calculate_readiness_score(
                close_series, analysis.get('ema20'),
                analysis.get('rsi'), analysis.get('adx'),
                volume, volume_20ma,
                analysis.get('days_above_ema20', 0),
                ema20_series=ema20_series,
                adx_series=adx_series,
                macd_histogram=analysis.get('macd_histogram')
            )

            # Applica smoothing + isteresi + jump override (anti-flickering)
            l2_smooth_state = self._apply_l2_smoothing(isin or ticker, l2_readiness_score)
            l2_smoothed_score = l2_smooth_state['smoothed']

            # Log con differenziazione based on transizioni
            if l2_smooth_state['enters_watchlist']:
                add_log(f"    🟨 L2 READINESS ENTRA: raw={l2_readiness_score:.0f} smoothed={l2_smoothed_score:.0f}")
            elif l2_smooth_state['exits_watchlist']:
                add_log(f"    ⬜ L2 READINESS ESCE: raw={l2_readiness_score:.0f} smoothed={l2_smoothed_score:.0f}")
            elif l2_smoothed_score >= 70:
                add_log(f"    🟨 L2 READINESS: raw={l2_readiness_score:.0f} smoothed={l2_smoothed_score:.0f} (watchlist)")

        except Exception as e:
            add_log(f"    ⚠️  L2 readiness score error: {e}")

        result = {
            'ticker':    ticker,
            'isin':      isin,
            'nome':      nome,
            'categoria': categoria,
            'borsa':     borsa,
            'livello':   level,
            'etf_type':  etf_type,
            'analysis':  analysis,
            'l0_signal': l0_entry_signal,
            'l1_tiered': l1_tiered,
            'l1_accelerated': l1_accelerated,  # STEP 12: Accelerated trigger
            'l0_regime_filter': l0_regime_filter,  # STEP 13: L0 regime
            'l1_seven_conditions': l1_seven_conditions,  # STEP 14: L1 7/7
            'l2_readiness_score': l2_readiness_score,  # STEP 15: L2 screening
        }

        return result

    def _empty_result(self, ticker, isin, nome, categoria, borsa, level, reason):
        famiglia = ETFTechnicalAnalyzer.detect_family(categoria)
        return {
            'ticker': ticker, 'isin': isin, 'nome': nome,
            'categoria': categoria, 'borsa': borsa, 'livello': level,
            'etf_type': famiglia,  # backward compatibility
            'analysis': {
                'current_price': None,
                'ema20': None, 'sma50': None, 'sma200': None,
                'rsi': None, 'adx': None, 'macd_histogram': None,
                'suggested_level': level, 'level_change': False,
                'level_reason': reason, 'conditions': {}, 'buy_count': 0,
                'l0_entry': False, 'l0_exit_rule': None,
                'pct_change_1d': None, 'pct_change_1w': None, 'pct_change_1m': None,
                'data_status': 'no_data',
            }
        }

    def update_excel(self, results: list):
        """Aggiorna il file Excel con i risultati."""
        try:
            wb = load_workbook(self.excel_path)
            ws = wb['ETF']

            COL_LIVELLO         = 1
            COL_TICKER          = 2
            COL_PREZZO          = 7
            COL_EMA20           = 8
            COL_SMA50           = 9
            COL_RSI             = 10
            COL_ADX             = 11
            COL_MACD            = 12
            COL_SEGNALE         = 13
            COL_ULTIMA_MODIFICA = 14

            excel_row_map = {row_idx - 2: row_idx for row_idx in range(2, ws.max_row + 1)}
            level_changes = []

            for i, result in enumerate(results):
                a = result['analysis']
                row = excel_row_map.get(i)
                if row is None:
                    continue

                current_level  = result['livello']
                suggested      = a.get('suggested_level', current_level)
                level_reason   = a.get('level_reason', '')

                if suggested != current_level:
                    ws.cell(row=row, column=COL_LIVELLO, value=suggested)
                    level_changes.append({
                        'nome': result['nome'], 'ticker': result['ticker'],
                        'isin': result.get('isin', ''),
                        'from': current_level, 'to': suggested,
                        'reason': level_reason,
                    })
                    cell = ws.cell(row=row, column=COL_LIVELLO)
                    if suggested < current_level:
                        cell.fill = PatternFill("solid", fgColor="00B050")
                        cell.font = Font(bold=True, color="FFFFFF")
                    else:
                        cell.fill = PatternFill("solid", fgColor="FF6600")
                        cell.font = Font(bold=True, color="FFFFFF")

                ws.cell(row=row, column=COL_PREZZO, value=a.get('current_price'))
                ws.cell(row=row, column=COL_EMA20,  value=a.get('ema20'))
                ws.cell(row=row, column=COL_SMA50,  value=a.get('sma50'))
                ws.cell(row=row, column=COL_RSI,    value=a.get('rsi'))
                ws.cell(row=row, column=COL_ADX,    value=a.get('adx'))
                ws.cell(row=row, column=COL_MACD,   value=a.get('macd_histogram'))

                # Segnale testuale
                if suggested == 0:
                    signal_txt = 'L0 RECOVERY'
                    color = 'FF6600'
                elif suggested == 1:
                    conds = a.get('buy_count', 0)
                    signal_txt = f'L1 BUY ({conds}/6)'
                    color = '00B050'
                elif suggested == 2:
                    signal_txt = 'L2 WATCH'
                    color = 'FFC000'
                else:
                    signal_txt = 'L3'
                    color = 'D9D9D9'

                cell = ws.cell(row=row, column=COL_SEGNALE, value=signal_txt)
                cell.fill = PatternFill("solid", fgColor=color)
                cell.font = Font(bold=True, color="000000" if color in ('FFC000', 'D9D9D9') else "FFFFFF")

                ws.cell(row=row, column=COL_ULTIMA_MODIFICA,
                        value=datetime.now().strftime('%Y-%m-%d %H:%M'))

            wb.save(self.excel_path)
            add_log(f"File Excel aggiornato")

            if level_changes:
                add_log(f"CAMBI DI LIVELLO: {len(level_changes)}")
                for c in level_changes:
                    dir_ = "UP" if c['to'] < c['from'] else "DOWN"
                    add_log(f"  {dir_} {c['nome'][:40]}: L{c['from']} → L{c['to']}")

            return level_changes

        except Exception as e:
            add_log(f"Errore aggiornamento Excel: {e}")
            add_log(traceback.format_exc())
            return []

    def generate_dashboard_data(self, results: list, send_daily_report: bool = True, errors: list = None) -> dict:
        """Genera dati JSON per la dashboard HTML."""
        if errors is None:
            errors = []
        l1_tracking = self.db.get_all_l1_entries()
        dashboard = {
            'last_update': datetime.now().isoformat(),
            'data_source': 'Yahoo Finance',
            'summary': {
                'total_etfs': len(results),
                'l0_count': 0, 'l1_count': 0, 'l2_count': 0, 'l3_count': 0,
                'alerts_sent': send_daily_report,
            },
            'levels': {0: [], 1: [], 2: [], 3: []},
            'categories': {},
        }

        for r in results:
            a         = r['analysis']
            suggested = int(a.get('suggested_level', r['livello']))  # Forza intero
            category  = r['categoria']

            level_key = f'l{suggested}_count'
            if level_key in dashboard['summary']:
                dashboard['summary'][level_key] += 1

            price = a.get('current_price')
            # Valutazione tiered e accelerated L1
            l1_tiered = r.get('l1_tiered', {})
            l1_accelerated = r.get('l1_accelerated', {})

            # Determina entry_mode (TIERED, ACCELERATED, NONE)
            entry_mode = 'NONE'
            if l1_accelerated.get('should_enter'):
                entry_mode = 'ACCELERATED'
            elif l1_tiered.get('should_enter'):
                entry_mode = 'TIERED'

            etf_data = {
                'ticker':            r['ticker'],
                'isin':              r.get('isin', ''),
                'nome':              r['nome'],
                'categoria':         category,
                'borsa':             r.get('borsa', ''),
                'etf_type':          r.get('etf_type', 'equity_developed'),
                'price':             float(price) if price is not None else None,
                'ema20':             a.get('ema20'),
                'sma50':             a.get('sma50'),
                'sma200':            a.get('sma200'),
                'rsi':               a.get('rsi'),
                'adx':               a.get('adx'),
                'regime':            a.get('regime', 'LATERALE'),  # NUOVO: regime per L0 sorting
                'macd_histogram':    a.get('macd_histogram'),
                'dist_ema20':        a.get('dist_ema20'),
                'days_above_ema20':  a.get('days_above_ema20', 0),
                'buy_count':         a.get('buy_count', 0),
                'level_reason':      a.get('level_reason', ''),
                'conditions':        a.get('conditions', {}),
                'pct_1d':            a.get('pct_change_1d'),
                'pct_1w':            a.get('pct_change_1w'),
                'pct_1m':            a.get('pct_change_1m'),
                'peak_price':        a.get('peak_price'),
                'drawdown_from_peak': a.get('drawdown_from_peak', 0.0),
                'l0_entry':          a.get('l0_entry', False),
                'data_status':       a.get('data_status', 'ok'),
                # STEP 10: Valutazione tiered L1
                'l1_tiered_entry':   l1_tiered.get('should_enter', False),
                'l1_confidence':     l1_tiered.get('confidence', 0.0),
                'l1_quality_score':  l1_tiered.get('quality_score', 0),
                'l1_quality_detail': l1_tiered.get('quality_detail', {}),
                'l1_tiered_reason':  l1_tiered.get('reason', ''),
                # STEP 12+: Valutazione accelerated L1 (gerarchia 2+2)
                'entry_mode':        entry_mode,
                'l1_accelerated_entry': l1_accelerated.get('should_enter', False),
                'l1_accelerated_confidence': l1_accelerated.get('confidence', 0.0),
                'l1_velocity_count': l1_accelerated.get('velocity_count', 0),
                'l1_velocity_detail': l1_accelerated.get('velocity_detail', {}),
                'l1_accelerated_reason': l1_accelerated.get('reason', ''),
            }

            # entry_date e entry_price per ETF in L1 (serve per linea verticale grafico)
            if suggested == 1:
                isin_key = r.get('isin', '') or r['ticker']
                tracking = l1_tracking.get(isin_key, {})
                etf_data['entry_date']  = str(tracking.get('entry_date', '')) if tracking else ''
                etf_data['entry_price'] = float(tracking.get('entry_price', 0)) if tracking else None

            level_int = max(0, min(3, int(suggested)))
            dashboard['levels'][level_int].append(etf_data)

            if category not in dashboard['categories']:
                dashboard['categories'][category] = []
            dashboard['categories'][category].append(etf_data)

        class SafeEncoder(json.JSONEncoder):
            def default(self, obj):
                if isinstance(obj, Decimal):
                    return float(obj)
                if isinstance(obj, np.integer):
                    return int(obj)
                if isinstance(obj, np.floating):
                    return float(obj)
                if isinstance(obj, np.ndarray):
                    return obj.tolist()
                if isinstance(obj, (date_type, datetime)):
                    return str(obj)
                return super().default(obj)

        # Data freshness: data più recente nel DB
        try:
            stats = self.db.get_stats()
            dashboard['summary']['data_as_of'] = stats.get('last_date')
        except Exception:
            dashboard['summary']['data_as_of'] = None

        # Health report per /api/health
        dashboard['health'] = {
            'timestamp': datetime.now().isoformat(),
            'total_etfs': len(results),
            'etfs_ok': len(results),
            'etfs_error': len(errors),
            'errors': [{'ticker': e.get('ticker', '?'), 'error': e.get('error', '')} for e in errors],
            'etfs_with_price': sum(1 for r in results if r['analysis'].get('current_price') is not None),
            'etfs_no_price': sum(1 for r in results if r['analysis'].get('current_price') is None),
        }

        os.makedirs('data', exist_ok=True)
        with open('data/dashboard_data.json', 'w') as f:
            json.dump(dashboard, f, indent=2, cls=SafeEncoder)

        return dashboard

    def update_trailing_stops(self, results: list):
        """
        Aggiorna gli stop loss trailing per le posizioni L1.
        Ogni giorno, per i profitti, il SL sale al 95% del prezzo corrente (ma non scende mai).
        """
        existing_l1 = self.db.get_all_l1_entries()

        for r in results:
            isin = r.get('isin', '') or r['ticker']
            if isin not in existing_l1:
                continue  # Non in L1

            a = r['analysis']
            current_price = a.get('current_price')
            entry_price = existing_l1[isin].get('entry_price')

            if not current_price or not entry_price:
                continue

            # Calcola guadagno % attuale
            pct_gain = (float(current_price) - float(entry_price)) / float(entry_price) * 100

            # Leggi config famiglia per SL
            fam_cfg = self.analyzer.p if hasattr(self, 'analyzer') else {}
            if not fam_cfg:
                # Fallback se analyzer non disponibile
                fam_cfg = {
                    'sl_profit_trigger_pct': 3.0,
                    'sl_trailing_tight_pct': 0.95
                }

            # Se in profitto e sopra soglia, attiva trailing
            profit_trigger = fam_cfg.get('sl_profit_trigger_pct', 3.0)
            if pct_gain >= profit_trigger:
                tight_pct = fam_cfg.get('sl_trailing_tight_pct', 0.95)
                sl_trailing = float(current_price) * tight_pct

                # Aggiorna nel DB (non scende mai)
                self.db.update_stop_loss_trailing(isin, sl_trailing)

                add_log(f"  Trailing SL: {r['nome'][:30]} — gain {pct_gain:.1f}% → SL={sl_trailing:.4f}")

    def send_alerts(self, results: list):
        """
        Invia alert giornalieri:
        - send_new_entries: solo nuovi ingressi in L1 (+ nuovi L0)
        - send_l1_exit: ogni uscita da L1
        - send_portfolio_signals: RSI > 72 o condizioni in deterioramento

        **IMPORTANTE**: usa results come source-of-truth per i livelli, NON il database
        (il DB potrebbe essere out-of-sync con i results appena calcolati)
        """
        today     = datetime.now().date()
        today_str = today.strftime('%Y-%m-%d')

        # Costruisci l'elenco dei livelli ATTUALI da results (source-of-truth)
        # Non leggiamo dal database per evitare out-of-sync con i results appena calcolati
        existing_l1 = {}
        existing_l0 = {}
        for r in results:
            isin = r.get('isin', '') or r['ticker']
            lvl = r['analysis'].get('suggested_level', r.get('livello', 3))
            if lvl == 1:
                existing_l1[isin] = r
            elif lvl == 0:
                existing_l0[isin] = r

        current_l1_isins  = set()
        current_l0_isins  = set()
        new_l1_entries    = []
        new_l0_entries    = []
        portfolio_signals = []

        for r in results:
            a         = r['analysis']
            suggested = a.get('suggested_level', r['livello'])
            isin      = r.get('isin', '') or r['ticker']
            price     = a.get('current_price')

            # ── L0 tracking ──────────────────────────────────────────────────
            if suggested == 0:
                current_l0_isins.add(isin)

                # PRIORITÀ 1 FASE 2 — Check invalidazione: prezzo < trigger_low_price
                l0_state = self.db.get_l0_state(isin)
                if l0_state and l0_state.get('trigger_low_price'):
                    trigger_low = l0_state['trigger_low_price']
                    if price and float(price) < trigger_low:
                        # Invalidazione: prezzo ha rotto il support → reset
                        self.db.remove_l0_entry(isin)
                        suggested = 3  # Ricade a L3
                        add_log(f"    ❌ L0 INVALIDAZIONE: prezzo {price:.2f} < trigger_low {trigger_low:.2f}")
                        continue  # Skip rest L0 processing

                # Fix 2026-08-05: rimossa la promozione automatica L0 → L2 su recovery_signal
                # confermato. Era la stessa logica della γ ora rimossa da suggest_level_0()
                # (vedi commento lì) applicata via un percorso diverso (etf_l0_tracking):
                # usava _get_l0_confirmation_signal — la stessa funzione richiesta per
                # CONFERMARE l'ingresso FAST/SLOW — come se fosse un segnale di USCITA.
                # L0 punta a inversioni di medio-lungo periodo: un ETF resta classificato
                # L0 finché non perde davvero i requisiti (β RSI<25, o invalidazione sotto
                # trigger_low_price qui sotto), non solo perché il recupero si conferma.

                # Estrai confirmation_mode e trigger_low_price per TUTTI i L0 (new o existing).
                # Fix 2026-08-05: prima leggeva da a.get('l0_regime_filter', {}).get('regime_type')
                # — un calcolo parallelo (l0_detect_regime_filter, solo informativo) con valori
                # 'fast_crash'/'slow_bear'/'none' mai uguali a 'FAST'/'SLOW' attesi altrove, e
                # comunque mai presente in 'a'. La fonte vera e' l0_data.l0_regime_mode, impostato
                # da suggest_level_0() quando il percorso FAST/SLOW determina davvero l'ingresso.
                confirmation_mode = a.get('l0_data', {}).get('l0_regime_mode')  # 'FAST' or 'SLOW' o None
                trigger_low_price = price  # Il prezzo al momento del trigger

                if isin not in existing_l0:
                    # NUOVO L0 entry
                    if price:
                        panic_low = a.get('l0_data', {}).get('panic_low')
                        self.db.set_l0_entry(isin, today_str, price, panic_low,
                                           confirmation_mode=confirmation_mode,
                                           trigger_low_price=trigger_low_price)
                    add_log(f"  NUOVO L0: {r['nome'][:40]} (mode={confirmation_mode})")
                    new_l0_entries.append({
                        'isin':              isin,
                        'ticker':            r['ticker'],
                        'nome':              r['nome'],
                        'price':             float(price) if price else None,
                        'panic_low':         a.get('l0_data', {}).get('panic_low'),
                        'rsi':               a.get('rsi'),
                        'distance_from_peak': a.get('l0_data', {}).get('distance_from_peak'),
                        'confirmation_mode': confirmation_mode,
                    })
                else:
                    # L0 entry ESISTENTE — aggiorna confirmation_mode se NULL
                    l0_state = self.db.get_l0_state(isin)
                    if l0_state and not l0_state.get('confirmation_mode'):
                        self.db.update_l0_state(isin, confirmation_mode=confirmation_mode,
                                               trigger_low_price=trigger_low_price)
                        add_log(f"  UPDATE L0: {r['nome'][:40]} (mode={confirmation_mode})")

                # Calcola e salva livello_display per tracking L0
                if price:
                    famiglia = ETFTechnicalAnalyzer.detect_family(r.get('categoria', ''))
                    analyzer = ETFTechnicalAnalyzer(famiglia=famiglia)
                    ema20 = a.get('ema20')
                    close_series = a.get('close_series')

                    livello_display = 'L0'
                    if ema20 is not None and float(price) > ema20:
                        livello_display = 'L2'
                        if close_series is not None:
                            sma50 = analyzer._sma(close_series.astype(float), 50)
                            sma50_v = analyzer._fval(sma50)
                            if sma50_v is not None and float(price) > sma50_v:
                                livello_display = 'L1'

                    self.db.update_l0_livello_display(isin, livello_display)

            # ── L1 tracking ──────────────────────────────────────────────────
            if suggested == 1:
                current_l1_isins.add(isin)
                # Controlla se è un NUOVO ingresso (non era in L1 nei results precedenti)
                if isin not in existing_l1:
                    if price:
                        sl_initial = a.get('stop_loss_initial')
                        # Salva SL iniziale (sarà poi accettato manualmente per diventare SL inserito)
                        self.db.set_l1_entry(isin, today_str, price, stop_loss=sl_initial)
                    entry_date  = today
                    entry_price = float(price) if price else None
                    sl_str = f" SL_sugg={a.get('stop_loss_initial'):.4f}" if a.get('stop_loss_initial') else ""
                    add_log(f"  NUOVO L1: {r['nome'][:40]}" +
                            (f" — entrato a {entry_price:.4f}{sl_str}" if entry_price else ''))
                    new_l1_entries.append({
                        'isin':      isin,
                        'ticker':    r['ticker'],
                        'nome':      r['nome'],
                        'categoria': r['categoria'],
                        'price':     float(price) if price else None,
                        'rsi':       a.get('rsi'),
                        'adx':       a.get('adx'),
                        'sma200':    a.get('sma200'),
                        'buy_count': a.get('buy_count', 0),
                    })
                else:
                    entry       = existing_l1[isin]
                    entry_date  = entry['entry_date']
                    entry_price = entry['entry_price']

                    try:
                        ed = entry_date if isinstance(entry_date, date_type) \
                            else datetime.fromisoformat(str(entry_date)).date()
                        days_in_l1 = max(1, int(np.busday_count(ed, today)) + 1)
                    except Exception:
                        days_in_l1 = 1

                    pct_gain = None
                    if price and entry_price:
                        pct_gain = round((float(price) - float(entry_price)) / float(entry_price) * 100, 2)

                    rsi = a.get('rsi') or 0
                    bc  = a.get('buy_count', 6)

                    signal_type   = None
                    signal_detail = None
                    if rsi >= 72:
                        signal_type   = 'stanchezza'
                        signal_detail = f'RSI attuale: {rsi:.1f} — zona di stanchezza in arrivo'
                    elif bc <= 4 and days_in_l1 > 5:
                        signal_type   = 'attenzione'
                        signal_detail = f'Condizioni soddisfatte: {bc}/6'

                    if signal_type:
                        portfolio_signals.append({
                            'isin':          isin,
                            'ticker':        r['ticker'],
                            'nome':          r['nome'],
                            'categoria':     r['categoria'],
                            'entry_date':    entry_date,
                            'days_in_l1':    days_in_l1,
                            'pct_gain':      pct_gain,
                            'rsi':           rsi,
                            'adx':           a.get('adx'),
                            'signal_type':   signal_type,
                            'signal_detail': signal_detail,
                        })

        # ── CONTROLLO STOP LOSS (Exit L1 se prezzo < SL) ───────────────────
        sl_triggered_exits = []
        for isin, entry in existing_l1.items():
            if isin in current_l1_isins:
                # ETF ancora in L1, ma verifica SL
                fr = next((r for r in results if (r.get('isin') or r['ticker']) == isin), None)
                if fr:
                    price = fr['analysis'].get('current_price')
                    sl_current = entry.get('stop_loss')

                    if price and sl_current and float(price) < float(sl_current):
                        # STOP LOSS TRIGGERATO!
                        sl_triggered_exits.append({
                            'isin': isin,
                            'ticker': fr.get('ticker'),
                            'nome': fr.get('nome'),
                            'entry_price': float(entry['entry_price']),
                            'entry_date': entry['entry_date'],
                            'exit_price': float(price),
                            'stop_loss': float(sl_current),
                            'pct_gain': round((float(price) - float(entry['entry_price'])) / float(entry['entry_price']) * 100, 2),
                        })
                        # Rimuovi da current_l1_isins per registrare come uscita
                        current_l1_isins.discard(isin)
                        add_log(f"  🚨 STOP LOSS TRIGGERATO: {fr['nome'][:40]} — Prezzo €{price:.4f} < SL €{sl_current:.4f}")

        # ── Uscite da L1 ──────────────────────────────────────────────────────
        for isin, entry in existing_l1.items():
            if isin in current_l1_isins:
                continue
            fr = next((r for r in results if (r.get('isin') or r['ticker']) == isin), None)
            if not fr:
                self.db.remove_l1_entry(isin)
                continue
            a          = fr['analysis']
            exit_price = a.get('current_price')
            ep         = entry['entry_price']
            ed_raw     = entry['entry_date']
            try:
                ed = ed_raw if isinstance(ed_raw, date_type) \
                    else datetime.fromisoformat(str(ed_raw)).date()
                days = max(1, int(np.busday_count(ed, today)) + 1)
            except Exception:
                days = 1
            pct = round((float(exit_price) - float(ep)) / float(ep) * 100, 2) \
                if exit_price and ep else None
            pct_str = f'{pct:+.2f}%' if pct is not None else 'N/D'
            add_log(f"  USCITA L1: {fr['nome'][:40]} — {pct_str}")
            # DISABILITATO: email uscite L1 (solo entrate + report portafoglio)
            # self.alert_system.send_l1_exit({
            #     'isin': isin, 'ticker': fr['ticker'], 'nome': fr['nome'],
            #     'categoria': fr['categoria'],
            #     'entry_date': ed_raw, 'entry_price': ep,
            #     'exit_price': float(exit_price) if exit_price else None,
            #     'days_in_l1': days, 'pct_gain': pct,
            #     'analysis': a,
            # })
            self.db.remove_l1_entry(isin)

        # ── Uscite da L0 ──────────────────────────────────────────────────────
        for isin in list(existing_l0.keys()):
            if isin not in current_l0_isins:
                add_log(f"  USCITA L0: {isin}")
                self.db.remove_l0_entry(isin)

        # ── INVIO ALERT STOP LOSS TRIGGERATI ────────────────────────────────
        for sl_exit in sl_triggered_exits:
            add_log(f"  📧 Alert SL: {sl_exit['nome'][:40]} — Registrato (no email)")
            # DISABILITATO: email uscite L1 per Stop Loss (solo entrate + report portafoglio)
            # self.alert_system.send_l1_exit({
            #     'isin': sl_exit['isin'],
            #     'ticker': sl_exit['ticker'],
            #     'nome': sl_exit['nome'],
            #     'categoria': '',
            #     'entry_date': sl_exit['entry_date'],
            #     'entry_price': sl_exit['entry_price'],
            #     'exit_price': sl_exit['exit_price'],
            #     'days_in_l1': 0,
            #     'pct_gain': sl_exit['pct_gain'],
            #     'analysis': {
            #         'current_price': sl_exit['exit_price'],
            #         'ema10': None,
            #         'ema20': None,
            #         'sma50': None,
            #         'rsi': None,
            #         'adx': None,
            #     },
            #     'exit_rule': 'Stop Loss',  # Identifica SL triggerato
            # })
            self.db.remove_l1_entry(sl_exit['isin'])

        # ── Invio email ────────────────────────────────────────────────────────
        if new_l1_entries or new_l0_entries:
            add_log(f"  Email nuovi ingressi: L1={len(new_l1_entries)} L0={len(new_l0_entries)}")
            self.alert_system.send_new_entries(new_l1_entries, new_l0_entries)
        else:
            add_log("  Nessun nuovo ingresso oggi")

        if portfolio_signals:
            add_log(f"  Email segnali portafoglio: {len(portfolio_signals)}")
            self.alert_system.send_portfolio_signals(portfolio_signals)
        else:
            add_log("  Nessun segnale portafoglio oggi")

    def run(self, send_daily_report: bool = True):
        """Esegue il ciclo completo di monitoraggio.

        Args:
            send_daily_report: Se True invia alert email (run principale).
                               Se False aggiorna solo dashboard (run silenzioso mattutino).
        """
        add_log("=" * 50)
        label = "completo" if send_daily_report else "silenzioso"
        add_log(f"ETF MONITOR [{label}] — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        add_log(f"Fonte dati: Yahoo Finance | Schema: EMA20+SMA50+SMA200+ADX+RSI")

        # 1. Carica ETF
        df_etfs = self.load_etfs()
        if df_etfs.empty:
            add_log("ERRORE: Nessun ETF da monitorare")
            return

        # 2. Analizza ogni ETF
        add_log(f"Analisi di {len(df_etfs)} ETF...")
        results = []
        errors  = []

        for idx, row in df_etfs.iterrows():
            try:
                result = self.analyze_etf(row)
                results.append(result)
                a = result['analysis']
                lvl = a.get('suggested_level', result['livello'])
                add_log(f"  OK L{lvl} | {row.get('Ticker','?'):12s} {row.get('Nome ETF','')[:28]}")
            except Exception as e:
                tb = traceback.format_exc()
                errors.append({'ticker': row.get('Ticker','?'), 'error': str(e)})
                add_log(f"  ERR {row.get('Ticker','?')}: {e}")
                add_log(f"  {tb}")
            time.sleep(self.data_fetcher.rate_limit)

        add_log(f"Analisi: {len(results)} OK, {len(errors)} errori")

        # 3. Aggiorna Excel
        try:
            add_log("Aggiornamento Excel...")
            self.update_excel(results)
        except Exception as e:
            add_log(f"ERRORE Excel: {e}")

        # 4. Genera dashboard
        try:
            add_log("Generazione dashboard...")
            dashboard = self.generate_dashboard_data(results, send_daily_report=send_daily_report, errors=errors)
            add_log(f"Dashboard: L0={dashboard['summary']['l0_count']} "
                    f"L1={dashboard['summary']['l1_count']} "
                    f"L2={dashboard['summary']['l2_count']} "
                    f"L3={dashboard['summary']['l3_count']}")
        except Exception as e:
            add_log(f"ERRORE Dashboard: {e}")
            add_log(traceback.format_exc())
            dashboard = {
                'last_update': datetime.now().isoformat() + 'Z',
                'data_source': 'Yahoo Finance',
                'summary': {'total_etfs': len(results),
                            'l0_count': 0, 'l1_count': 0, 'l2_count': 0, 'l3_count': 0},
                'levels': {0: [], 1: [], 2: [], 3: []},
                'categories': {}
            }
            for r in results:
                try:
                    lvl = r['analysis'].get('suggested_level', r['livello'])
                    dashboard['levels'][max(0, min(3, int(lvl)))].append({
                        'ticker': r['ticker'], 'nome': r['nome'],
                        'price':  r['analysis'].get('current_price'),
                    })
                except Exception:
                    pass
            with open('data/dashboard_data.json', 'w') as f:
                json.dump(dashboard, f, indent=2)

        # PRIORITÀ 3 — Salva stato L2 smoothing per anti-flickering (prossimo ciclo)
        try:
            self._save_l2_state()
            add_log("✓ L2 state salvato (smoothing/isteresi)")
        except Exception as e:
            add_log(f"⚠️  Errore salvataggio L2 state: {e}")

        # 5. Aggiorna trailing stops (daily)
        try:
            add_log("Aggiornamento trailing stops...")
            self.update_trailing_stops(results)
        except Exception as e:
            add_log(f"ERRORE Trailing stops: {e}")

        # 6. Invia alert (solo se run principale)
        if send_daily_report:
            try:
                add_log("Invio alert...")
                self.send_alerts(results)
            except Exception as e:
                add_log(f"ERRORE Alert: {e}")
        else:
            add_log("Alert saltati (refresh silenzioso)")

        # STEP 4 — Aggiorna SL/SG suggerito per portafoglio L1
        tightening_events = []
        try:
            add_log("Aggiornamento SL/SG suggerito L1...")
            tightening_events += self._update_portfolio_l1_suggerito(results) or []
        except Exception as e:
            add_log(f"⚠️  Errore aggiornamento L1 suggerito: {e}")
            add_log(traceback.format_exc())

        # STEP 7 — Aggiorna SL/TP suggerito per portafoglio L0
        try:
            add_log("Aggiornamento SL/TP suggerito L0...")
            tightening_events += self._update_portfolio_l0_suggerito(results) or []
        except Exception as e:
            add_log(f"⚠️  Errore aggiornamento L0 suggerito: {e}")
            add_log(traceback.format_exc())

        # STEP 7b — Alert dedicato "avvicinamento TP" (2026-08-20): invia SUBITO,
        # non aspetta il resoconto serale (send_daily_report) — questo è il punto,
        # arrivare prima del giro schedulato successivo. Gira anche sul run
        # silenzioso del mattino, a differenza del resto degli alert sopra.
        if tightening_events:
            try:
                add_log(f"Invio alert avvicinamento TP ({len(tightening_events)} posizioni)...")
                self.alert_system.send_tp_proximity_alert(tightening_events)
            except Exception as e:
                add_log(f"⚠️  Errore alert avvicinamento TP: {e}")

        # STEP 8 — Shadow Monitor CANDIDATE_MODEL_B_20260807 (2026-08-07): traccia
        # posizioni ipotetiche sul cluster 'core' con i parametri del candidato
        # (mm200_distance_max=7.0%, adx-4, smart_6_macd, TP=15%) senza toccare NESSUNA
        # decisione reale — solo log su etf_shadow_positions per il confronto
        # native_7 vs candidato a fine lockdown (06/09/2026). Email sui nuovi ingressi
        # collegata il 2026-08-19 su richiesta esplicita (prima "nessuna email").
        # Vedi CLAUDE.md "CANDIDATE_MODEL_B_20260807".
        try:
            from shadow_monitor import run_shadow_monitor
            shadow_l1_entries = run_shadow_monitor(self.db, results, add_log=add_log)
            if shadow_l1_entries and send_daily_report:
                self.alert_system.send_shadow_entries(shadow_l1_entries, variant='L1')
        except Exception as e:
            add_log(f"⚠️  Errore Shadow Monitor (non bloccante): {e}")

        # STEP 8b — Shadow Monitor CANDIDATE_MODEL_L0_20260808 (2026-08-08): stesso
        # principio dello Shadow Monitor L1 sopra, ma per L0 — traccia posizioni
        # ipotetiche su equity_sviluppati (unica famiglia raggiungibile da L0) con
        # regime_min_days_below_sma200=5 invece del baseline YAML=10, senza toccare
        # NESSUNA decisione reale. Solo log su etf_shadow_positions per il confronto
        # a fine lockdown (06/09/2026). Email sui nuovi ingressi collegata il
        # 2026-08-19 su richiesta esplicita (prima "nessuna email"). Vedi CLAUDE.md
        # "CANDIDATE_MODEL_L0_20260808".
        try:
            from shadow_monitor_l0 import run_shadow_monitor_l0
            shadow_l0_entries = run_shadow_monitor_l0(self.db, results, add_log=add_log)
            if shadow_l0_entries and send_daily_report:
                self.alert_system.send_shadow_entries(shadow_l0_entries, variant='L0')
        except Exception as e:
            add_log(f"⚠️  Errore Shadow Monitor L0 (non bloccante): {e}")

        # STEP 8c — Shadow Monitor "terza via" Market Breadth (2026-08-20): traccia
        # SOLO i trade extra che il gate 6/7+MACD aprirebbe rispetto al sistema reale
        # (native_7), e SOLO quando la breadth di mercato (EMA20>SMA50 su tutto
        # l'universo) e' >=80% (isteresi: esce sotto 65%). Nessuna decisione reale
        # toccata. Vedi memory/etf_post_lockdown_todo_20260906.md sezione 3 per il
        # backtest e lo sweep di soglie che l'hanno preceduto.
        try:
            from shadow_monitor_breadth import run_shadow_monitor_breadth
            shadow_breadth_entries = run_shadow_monitor_breadth(self.db, results, add_log=add_log)
            if shadow_breadth_entries and send_daily_report:
                self.alert_system.send_shadow_entries(shadow_breadth_entries, variant='BREADTH')
        except Exception as e:
            add_log(f"⚠️  Errore Shadow Monitor Breadth (non bloccante): {e}")

        # (Shadow Monitor CANDIDATE_MODEL_L0_SL_20260820 rimosso il 2026-08-20 lo
        # stesso giorno in cui e' stato creato: il candidato [primo scaglione SL L0
        # 2%->4%] e' stato promosso direttamente in produzione su richiesta esplicita
        # dell'utente [technical_analysis.py::calculate_sl_suggerito_l0], quindi non
        # esiste piu' una "baseline" distinta da tracciare in parallelo — produzione
        # e candidato ora coincidono. shadow_monitor_l0_sl.py eliminato dal repo,
        # le 2 posizioni ombra aperte quel giorno chiuse amministrativamente nel DB
        # (exit_reason='PROMOTED'). Vedi CLAUDE.md e
        # memory/etf_post_lockdown_todo_20260906.md sezione 3.)

        # STEP 8d — Shadow Monitor CANDIDATE_L0_ORO_20260824: traccia posizioni
        # ipotetiche di L0 (mean-reversion) su oro_metalli_preziosi, whitelist L0
        # bypassata SOLO in memoria dentro shadow_monitor_l0_oro.py (mai scritta su
        # YAML) — nessuna decisione reale toccata. Nato da un backtest 2026-08-24 che
        # ha mostrato L0 promettente sull'oro (N=3, WR 66.7%, +2.096€ netti) ma con
        # campione troppo piccolo per essere conclusivo — questo Shadow Monitor
        # accumula dati forward reali prima di decidere se aprire davvero la
        # whitelist. Vedi memory/etf_family_viability_survey_2026_08_24.md.
        try:
            from shadow_monitor_l0_oro import run_shadow_monitor_l0_oro
            shadow_l0_oro_entries = run_shadow_monitor_l0_oro(self.db, results, add_log=add_log)
            if shadow_l0_oro_entries and send_daily_report:
                self.alert_system.send_shadow_entries(shadow_l0_oro_entries, variant='L0_ORO')
        except Exception as e:
            add_log(f"⚠️  Errore Shadow Monitor L0-oro (non bloccante): {e}")

        # STEP 8e — Shadow Monitor CANDIDATE_L0_METALLI_20260824: stesso principio
        # dello Shadow Monitor L0-oro sopra, ma per metalli_industriali — whitelist L0
        # bypassata SOLO in memoria dentro shadow_monitor_l0_metalli.py (mai scritta su
        # YAML). Backtest 2026-08-24: N=13 (piu' solido dell'oro), WR 53.8%, +5.088€
        # netti. Vedi memory/etf_family_viability_survey_2026_08_24.md.
        try:
            from shadow_monitor_l0_metalli import run_shadow_monitor_l0_metalli
            shadow_l0_metalli_entries = run_shadow_monitor_l0_metalli(self.db, results, add_log=add_log)
            if shadow_l0_metalli_entries and send_daily_report:
                self.alert_system.send_shadow_entries(shadow_l0_metalli_entries, variant='L0_METALLI')
        except Exception as e:
            add_log(f"⚠️  Errore Shadow Monitor L0-metalli (non bloccante): {e}")

        # 7. Invia resoconto portafoglio — DOPO l'aggiornamento SL/TP (fix 2026-08-05:
        # prima veniva inviato PRIMA degli STEP 4/7 sopra, quindi l'email mostrava
        # sempre i valori SL/TP del giorno precedente invece di quelli appena calcolati
        # sul prezzo di oggi)
        if send_daily_report:
            favorites_digest = []
            try:
                favorites_digest = self._build_favorites_digest(results)
            except Exception as e:
                add_log(f"⚠️  Errore digest Preferiti: {e}")
            l2_radar = []
            try:
                l2_radar = self._build_l2_radar(results)
            except Exception as e:
                add_log(f"⚠️  Errore Radar L2: {e}")
            try:
                add_log("Invio resoconto portafoglio...")
                self.alert_system.send_portfolio_report(favorites_digest=favorites_digest, l2_radar=l2_radar)
            except Exception as e:
                add_log(f"ERRORE Resoconto portafoglio: {e}")

        # STEP 8 — Sincronizza segnali L1 al portafoglio personale
        try:
            from sync_l1_portfolio import PortfolioL1Syncer
            
            # Filtra i segnali L1 (suggested_level == 1)
            l1_signals = [r for r in results if r.get('suggested_level') == 1]
            
            if l1_signals:
                syncer = PortfolioL1Syncer(self.db)
                sync_result = syncer.sync_l1_entries(l1_signals)
                add_log(f"Portfolio L1 Sync: {sync_result['added']} added, {sync_result['skipped']} skipped")
                
                # Invia email con nuove posizioni aggiunte
                if sync_result['added'] > 0:
                    for detail in sync_result['details']:
                        if detail['added']:
                            add_log(f"  ✅ {detail['ticker']} → Portfolio L1 (Entry €{detail['entry_price']:.2f})")
            else:
                add_log("Portfolio L1 Sync: No new L1 signals to add")
                
        except Exception as e:
            add_log(f"⚠️  Portfolio L1 Sync error: {e}")


        # STEP 9 — Prepara candidati L0 per email
        try:
            add_log("Raccolta candidati L0...")
            l0_candidates = self._collect_l0_candidates(results)
            if l0_candidates:
                add_log(f"  Trovati {len(l0_candidates)} segnali L0 (alert-only)")
        except Exception as e:
            add_log(f"⚠️  Errore raccolta L0 candidati: {e}")

        # Aggiorna stop loss suggerito per portafoglio (formula continua)
        try:
            self._update_portfolio_sl_suggested()
        except Exception as e:
            add_log(f"⚠️  Errore aggiornamento portfolio SL: {e}")

        # STEP 10 — Rigenera PDF (sincronizzati con YAML)
        try:
            generate_parameters_pdf('data/ETF_Monitor_Parametri_Riferimento.pdf')
            generate_complete_pdf('data/ETF_Monitor_Documentazione_Completa.pdf')
            add_log("✅ PDF rigenerati (parametri + documentazione completa — sincronizzati con YAML)")
        except Exception as e:
            add_log(f"⚠️  Errore generazione PDF: {e}")

        add_log(f"Completato — {datetime.now().strftime('%H:%M')}")
        add_log("=" * 50)


    def _collect_l0_candidates(self, results: list) -> list:
        """
        STEP 9 — Raccoglie gli ETF con segnali L0 candidati (alert-only).

        Riporta gli ETF che hanno segnali L0 validi ma non sono ancora
        in portafoglio. Usato per suggerire nuovi ingressi L0.

        Returns:
            Lista di dict con candidati L0
        """
        l0_candidates = []

        try:
            # Leggi ETF già in L0 portafoglio
            conn = self.db.get_connection()
            if not conn:
                return []

            with conn.cursor() as cur:
                cur.execute("""
                    SELECT DISTINCT isin FROM etf_portfolio_entries
                    WHERE portafoglio = 'L0' AND status = 'active'
                """)
                existing_l0 = {row[0] for row in cur.fetchall()}
            conn.close()

        except Exception as e:
            add_log(f"  ⚠️  Errore lettura L0 esistenti: {e}")
            existing_l0 = set()

        # Filtra candidati L0 dai risultati dell'analisi
        for result in results:
            # Salta se già in L0 portafoglio
            if result.get('isin') in existing_l0:
                continue

            l0_signal = result.get('l0_signal', {})
            if not l0_signal.get('l0_signal'):
                continue

            # È un candidato L0
            isin = result.get('isin')
            ticker = result.get('ticker')
            price = result.get('analysis', {}).get('current_price')

            # DEBUG: log candidati L0
            add_log(f"  L0 CANDIDATO: {ticker} | ISIN: {isin} | Price: {price} | DD: {l0_signal.get('dd_from_high'):.1f}%")

            # SALVA il segnale L0 nel DB (se non esiste già)
            if isin and price:
                self.db.set_l0_entry(isin, datetime.now().strftime('%Y-%m-%d'), float(price))
                add_log(f"    ✓ Salvato in etf_l0_tracking")

            candidate = {
                'ticker': result.get('ticker'),
                'isin': isin,
                'nome': result.get('nome'),
                'categoria': result.get('categoria'),
                'price': price,
                'dd_from_high': l0_signal.get('dd_from_high'),
                'rsi': l0_signal.get('rsi'),
                'recovery': l0_signal.get('recovery'),
                'alert_only': l0_signal.get('alert_only', True),
            }
            l0_candidates.append(candidate)

        return l0_candidates

    def _update_portfolio_l0_suggerito(self, results: list) -> list:
        """
        STEP 7 — Aggiorna SL/TP suggeriti per posizioni L0 attive.

        Ritorna la lista dei nuovi eventi di "stringimento tattico" (Stop che
        entra o avanza nella fascia di avvicinamento al TP, vedi
        order_pricing.py) rilevati in QUESTO giro — usata da monitor.py per
        l'alert email dedicato (2026-08-20, non aspettare il resoconto serale).

        Fix 2026-08-05: prima l'uscita reale passava da check_l0_exit() (kill switch,
        bear trap RSI<25, stop assoluto min30gg, timeout 45gg) — la stessa contraddizione
        già corretta su L1: il sistema non esegue mai ordini in automatico. Sostituito
        allora con un tocco di SL/TP che marcava 'exited' in automatico nel DB — una
        contraddizione più sottile ma identica nella sostanza (il prezzo di chiusura
        calcolato qui non è mai garanzia del prezzo reale eseguito su Directa). Fix
        2026-08-22: rimosso anche questo automatismo, vedi punto 4. La posizione NON
        viene mai riclassificata a L1/L2: resta in portafoglio L0 finché l'utente non
        conferma manualmente l'uscita reale, coerente con la logica "medio-lungo
        periodo" di L0.

        Per ogni entry in portafoglio L0:
        1. Recupera dati di mercato (price, ema20)
        2. Chiama calculate_sl_suggerito_l0() → SL trailing progressivo
        3. Chiama calculate_tp_suggerito_l0() → TP fisso di famiglia
        4. Se il prezzo tocca SL o TP: SOLO log/alert informativo (mai 'exited'
           automatico — l'unica uscita reale è la conferma manuale dell'utente in
           dashboard col prezzo davvero eseguito su Directa)
        5. Sempre: salva SL/TP e contatori aggiornati (posizione resta 'active')
        """
        tightening_events = []
        try:
            # Leggi tutte le entry L0 attive
            query = """
                SELECT id, isin, entry_date, entry_price, fund_name,
                       days_no_recovery, stallo_counter, broker, tp_proximity_stop_max
                FROM etf_portfolio_entries
                WHERE status = 'active' AND portafoglio = 'L0'
            """
            conn = self.db.get_connection()
            if not conn:
                add_log("  ⚠️  Nessuna connessione DB")
                return tightening_events

            try:
                with conn.cursor() as cur:
                    cur.execute(query)
                    rows = cur.fetchall()
            except Exception as e:
                add_log(f"  ⚠️  Errore query portfolio L0: {e}")
                conn.close()
                return tightening_events

            if not rows:
                add_log("  — Nessuna posizione L0 da aggiornare")
                conn.close()
                return tightening_events

            add_log(f"  Aggiornamento {len(rows)} posizioni L0...")

            for entry_id, isin, entry_date_str, entry_price_str, fund_name, days_no_rec, stallo_cnt, broker, prev_tp_stop_max in rows:
                try:
                    entry_price = float(entry_price_str) if entry_price_str else None
                    if not entry_price or entry_price <= 0:
                        continue

                    # Trova il risultato corrispondente per recuperare i dati di mercato
                    result = None
                    for r in results:
                        if (r.get('isin') == isin or r.get('ticker') == isin):
                            result = r
                            break

                    if not result or not result.get('analysis'):
                        continue

                    a = result['analysis']
                    current_price = a.get('current_price')
                    if not current_price:
                        continue

                    current_price = float(current_price)

                    # Recupera famiglia per i parametri L0
                    famiglia = ETFTechnicalAnalyzer.detect_family(fund_name or result.get('categoria', ''))
                    analyzer = ETFTechnicalAnalyzer(famiglia=famiglia)
                    sl_initial_pct = analyzer.p.get('sl_initial_pct')

                    # Dati di mercato per SL
                    ema20 = a.get('ema20')

                    # ATR reale dello strumento (non della famiglia) — usato dal ratchet di
                    # avvicinamento al TP (order_pricing.py) per scegliere il buffer stretto/
                    # largo. Motivato dal caso reale PHAG/WisdomTree Physical Silver
                    # (2026-08-22): la famiglia oro_metalli_preziosi ha sl_initial_pct=5%
                    # (sotto la soglia 7% del buffer largo), ma raggruppa Oro e Argento —
                    # l'Argento è storicamente molto più volatile del proxy di famiglia
                    # (ATR14 misurato 2,87% contro un buffer "stretto" pensato per l'1%,
                    # rischio concreto di whipsaw). L'ATR per-strumento supera questo limite
                    # senza dover spostare l'intera famiglia (che include anche l'Oro, molto
                    # meno volatile) su un profilo di rischio diverso.
                    atr_pct = None
                    try:
                        hist_atr = self.db.get_ohlc_by_isin(isin, days=40)
                        if not hist_atr.empty and len(hist_atr) >= 14:
                            atr_norm = analyzer._calculate_atr_normalized(
                                hist_atr['High'], hist_atr['Low'], hist_atr['Close'].astype(float))
                            if atr_norm is not None:
                                atr_pct = round(atr_norm * 100, 2)
                    except Exception as e:
                        add_log(f"    ⚠️  Errore ATR {isin}: {e}")

                    # CALCOLA SL SUGGERITO — trailing progressivo
                    sl_data = analyzer.calculate_sl_suggerito_l0(entry_price, current_price)
                    sl_suggerito = sl_data.get('sl_suggerito')
                    stage = sl_data.get('stage')

                    # CALCOLA TP SUGGERITO — target fisso di famiglia
                    tp_data = analyzer.calculate_tp_suggerito_l0(entry_price, current_price)
                    tp_suggerito = tp_data.get('tp_suggerito')

                    profit_pct = (current_price - entry_price) / entry_price if entry_price > 0 else 0

                    # Contatore stallo: se profit è tra -1% e +2%
                    if -0.01 <= profit_pct <= 0.02:
                        stallo_cnt = (stallo_cnt or 0) + 1
                    else:
                        stallo_cnt = 0

                    # Contatore days_no_recovery: se price < EMA20
                    if ema20 is not None and current_price < ema20:
                        days_no_rec = (days_no_rec or 0) + 1
                    else:
                        days_no_rec = 0

                    # SEGNALI INFORMATIVI — SL/TP raggiunti (SOLO log/alert, MAI un
                    # automatismo di chiusura: il sistema non esegue mai ordini, l'unica
                    # uscita reale è la conferma manuale dell'utente in dashboard
                    # (/api/portfolio/<isin>/exit) col prezzo davvero eseguito su
                    # Directa — vedi CLAUDE.md "Come opera davvero l'utente". Prima
                    # questo blocco marcava 'exited' in automatico usando il prezzo di
                    # chiusura calcolato qui, assumendo che coincidesse col fill reale:
                    # esattamente la contraddizione già segnalata e corretta su L1 il
                    # 2026-08-05, qui però era ancora presente. Fix 2026-08-22.
                    sl_hit = sl_suggerito is not None and current_price <= sl_suggerito
                    tp_hit = tp_data.get('trigger', False)

                    # Ratchet dello Stop tattico di avvicinamento al TP (order_pricing.py)
                    # — gira SEMPRE, anche oltre il TP: dato che non si esce mai da soli,
                    # lo Stop deve restare stretto a ridosso/oltre il target finché
                    # l'utente non conferma l'uscita reale.
                    op = compute_order_prices(
                        current_price, sl_suggerito, tp_suggerito, broker,
                        previous_tightened_stop=float(prev_tp_stop_max) if prev_tp_stop_max else None,
                        sl_initial_pct=sl_initial_pct,
                        atr_pct=atr_pct,
                    )
                    tp_proximity_stop_max = op.get('tp_proximity_stop_max')
                    if op.get('tightened'):
                        add_log(f"    🔶 Stop tattico stretto a {op['prezzo_stop']:.2f}€ (avvicinamento/superamento TP)")
                        prev_val = float(prev_tp_stop_max) if prev_tp_stop_max else None
                        if prev_val is None or abs(tp_proximity_stop_max - prev_val) > 0.0001:
                            tightening_events.append({
                                'isin': isin, 'ticker': result.get('ticker'),
                                'fund_name': fund_name, 'variant': 'L0',
                                'current_price': current_price,
                                'prezzo_stop': op.get('prezzo_stop'),
                                'prezzo_limite_stop': op.get('prezzo_limite_stop'),
                                'tp_suggerito': tp_suggerito, 'broker': broker,
                            })

                    # Salva SL, TP e contatori — MAI 'exited' automaticamente
                    with conn.cursor() as cur:
                        cur.execute("""
                            UPDATE etf_portfolio_entries
                            SET sl_suggerito = %s, sg_suggerito = %s, days_no_recovery = %s,
                                stallo_counter = %s, stop_loss_updated_at = now(),
                                tp_proximity_stop_max = %s
                            WHERE id = %s
                        """, (sl_suggerito, tp_suggerito, days_no_rec, stallo_cnt, tp_proximity_stop_max, entry_id))
                        conn.commit()

                    alert_msg = ''
                    if tp_hit:
                        alert_msg = ' 🟢 TARGET TP RAGGIUNTO — conferma uscita reale manualmente in dashboard'
                    elif sl_hit:
                        alert_msg = ' 🔴 SL RAGGIUNTO — conferma uscita reale manualmente in dashboard'
                    elif stallo_cnt and stallo_cnt >= 20:
                        alert_msg = f' ⚠️ STALLO {stallo_cnt}gg'
                    elif days_no_rec and days_no_rec >= 40:
                        alert_msg = f' ⏱️ TIMEOUT {days_no_rec}gg'

                    add_log(f"    {fund_name[:40]:40s} | SL: {sl_suggerito:.2f}€ ({stage}) | TP: {tp_suggerito:.2f}€{alert_msg}")

                except Exception as e:
                    add_log(f"    ⚠️  Errore L0 {isin}: {type(e).__name__}: {e}")
                    continue

            conn.close()

        except Exception as e:
            add_log(f"  ⚠️  Errore generale L0: {e}")

        return tightening_events

    def _build_favorites_digest(self, results: list) -> list:
        """
        Costruisce il digest 'Preferiti' per l'email quotidiana: per ogni ETF
        seguito manualmente (etf_favorites, lista curata dall'utente — non va
        confusa con etf_l2_watchlist, lo stato interno di anti-flickering del
        punteggio L2), legge lo stato di oggi dai risultati appena calcolati,
        lo confronta con l'ultimo stato salvato (last_buy_count/last_level/
        last_regime) per i delta, poi aggiorna lo snapshot per il confronto
        di domani.
        """
        digest = []
        favorites = self.db.get_favorites()
        if not favorites:
            return digest

        results_by_key = {(r.get('isin', '') or r['ticker']): r for r in results}

        for fav in favorites:
            isin = fav['isin']
            r = results_by_key.get(isin)
            if not r:
                continue
            a = r.get('analysis', {})
            buy_count = a.get('buy_count', 0)
            level     = int(a.get('suggested_level', r.get('livello', 3)))
            regime    = a.get('regime', 'LATERALE')

            last_buy_count = fav.get('last_buy_count')
            last_level     = fav.get('last_level')
            last_regime    = fav.get('last_regime')

            digest.append({
                'isin':             isin,
                'ticker':           r['ticker'],
                'nome':             r['nome'],
                'buy_count':        buy_count,
                'level':            level,
                'regime':           regime,
                'delta_buy_count':  (buy_count - last_buy_count) if last_buy_count is not None else None,
                'level_changed':    last_level is not None and level != last_level,
                'regime_changed':   last_regime is not None and regime != last_regime,
            })

            self.db.update_favorite_snapshot(isin, buy_count, level, regime)

        return digest

    # Etichette leggibili delle 7 condizioni L1 (stesso ordine/nomi di CLAUDE.md
    # "L1 — Come Si Entra"), usate per tradurre il dict 'conditions' in un elenco
    # "cosa manca" nel Radar L2.
    _L1_CONDITION_LABELS = {
        'allineamento_ok':  'Allineamento',
        'persistenza_ok':   'Persistenza',
        'rsi_ok':           'RSI',
        'distance_ok':      'Distanza EMA20',
        'adx_ok':           'ADX',
        'macd_ok':          'MACD',
        'space_residuo_ok': 'Spazio Residuo',
    }

    def _build_l2_radar(self, results: list) -> list:
        """
        Costruisce il 'Radar L2' per l'email quotidiana: a differenza del digest
        Preferiti (lista curata a mano dall'utente), questo scandisce TUTTO
        l'universo classificato L2 oggi (buy_count>=5/7, vedi
        technical_analysis.py::suggest_level()) e per ognuno mostra quale
        condizione manca ancora per L1 — non solo il conteggio, come fa invece
        il tab dashboard.

        Include anche il caso "7/7 ma bloccato dalle fondamenta" (regime non
        BULL o prezzo<SMA50, vedi CLAUDE.md "FONDAMENTA IRRINUNCIABILI") — qui
        le 7 condizioni sono tutte vere, quindi 'conditions' non aiuta: il
        motivo si legge da 'level_reason' (livello 2, buy_count già a 7).
        """
        radar = []
        for r in results:
            a = r.get('analysis', {}) or {}
            if a.get('suggested_level') != 2:
                continue

            buy_count = a.get('buy_count', 0)
            conditions = a.get('conditions', {}) or {}
            level_reason = a.get('level_reason', '') or ''

            if buy_count >= 7:
                if 'non BULL' in level_reason:
                    missing = ['Regime non BULL']
                elif '< SMA50' in level_reason:
                    missing = ['Prezzo < SMA50']
                else:
                    missing = ['Fondamenta']
            else:
                missing = [label for key, label in self._L1_CONDITION_LABELS.items()
                           if conditions.get(key) is False]
                if not missing:
                    missing = ['—']

            radar.append({
                'isin':      r.get('isin'),
                'ticker':    r.get('ticker'),
                'nome':      r.get('nome'),
                'buy_count': buy_count,
                'missing':   missing,
                'regime':    conditions.get('regime') or a.get('regime'),
            })

        radar.sort(key=lambda x: (-x['buy_count'], x['ticker'] or ''))
        return radar[:20]

    def _update_portfolio_l1_suggerito(self, results: list) -> list:
        """
        STEP 4 — Aggiorna SL/SG suggerito per posizioni L1 attive.

        Per ogni entry in portafoglio L1 (fix 2026-08-05, modello reale confermato;
        fix 2026-08-22: rimosso l'auto-exit, vedi punto 4):
        1. Recupera dati di mercato (price, ema20, ema20_series, etc)
        2. Chiama calculate_sl_suggerito_l1() → SL
        3. Chiama calculate_stop_gain_dynamic() → TP (unica funzione TP, niente decadimento
           temporale, si muove con lo slope EMA20)
        4. Se il prezzo tocca SL o scatta il trigger del TP: SOLO log/alert informativo
           (mai 'exited' automatico — il sistema non esegue mai ordini, l'unica uscita
           reale è la conferma manuale dell'utente in dashboard col prezzo davvero
           eseguito su Directa)
        5. Sempre: salva sl_suggerito/sg_suggerito aggiornati (posizione resta 'active'
           finché l'utente non conferma l'uscita reale)

        Ritorna la lista dei nuovi eventi di "stringimento tattico" rilevati in
        QUESTO giro — stesso meccanismo di _update_portfolio_l0_suggerito, vedi lì.
        """
        tightening_events = []
        try:
            # Leggi tutte le entry L1 attive
            query = """
                SELECT id, isin, entry_date, entry_price, fund_name,
                       broker, tp_proximity_stop_max
                FROM etf_portfolio_entries
                WHERE status = 'active' AND portafoglio = 'L1'
            """
            conn = self.db.get_connection()
            if not conn:
                add_log("  ⚠️  Nessuna connessione DB")
                return tightening_events

            try:
                with conn.cursor() as cur:
                    cur.execute(query)
                    rows = cur.fetchall()
            except Exception as e:
                add_log(f"  ⚠️  Errore query portfolio L1: {e}")
                conn.close()
                return tightening_events

            if not rows:
                add_log("  — Nessuna posizione L1 da aggiornare")
                conn.close()
                return tightening_events

            add_log(f"  Aggiornamento {len(rows)} posizioni L1...")

            for entry_id, isin, entry_date_str, entry_price_str, fund_name, broker, prev_tp_stop_max in rows:
                try:
                    entry_price = float(entry_price_str) if entry_price_str else None
                    if not entry_price or entry_price <= 0:
                        continue

                    # Trova il risultato corrispondente per recuperare i dati di mercato
                    result = None
                    for r in results:
                        if (r.get('isin') == isin or r.get('ticker') == isin):
                            result = r
                            break

                    if not result or not result.get('analysis'):
                        continue

                    a = result['analysis']
                    current_price = a.get('current_price')
                    if not current_price:
                        continue

                    current_price = float(current_price)

                    # Recupera famiglia dall'analyzer per i parametri
                    famiglia = ETFTechnicalAnalyzer.detect_family(fund_name or result.get('categoria', ''))
                    analyzer = ETFTechnicalAnalyzer(famiglia=famiglia)
                    sl_initial_pct = analyzer.p.get('sl_initial_pct')

                    # Dati di mercato per SL/SG
                    ema20 = a.get('ema20')

                    # Serie EMA20 per lo slope di calculate_stop_gain_dynamic() — 'a' (analysis)
                    # non la contiene mai (analyze_etf() non imposta quella chiave), quindi la
                    # ricaviamo da uno storico Close indipendente via get_ohlc_by_isin
                    # (fast-path DB, gia' aggiornato in questo stesso run).
                    # ATR reale dello strumento (non della famiglia) — vedi commento gemello
                    # in _update_portfolio_l0_suggerito per il caso reale che l'ha motivato
                    # (PHAG/Silver, 2026-08-22). Riusa lo stesso hist_recent già scaricato
                    # sopra per l'EMA20 series.
                    ema20_series = None
                    atr_pct = None
                    try:
                        hist_recent = self.db.get_ohlc_by_isin(isin, days=40)
                        if not hist_recent.empty and len(hist_recent) >= 20:
                            close_recent = hist_recent['Close'].astype(float).dropna()
                            if len(close_recent) >= 20:
                                ema20_series = analyzer._ema(close_recent, 20).tail(10)
                        if not hist_recent.empty and len(hist_recent) >= 14:
                            atr_norm = analyzer._calculate_atr_normalized(
                                hist_recent['High'], hist_recent['Low'], hist_recent['Close'].astype(float))
                            if atr_norm is not None:
                                atr_pct = round(atr_norm * 100, 2)
                    except Exception as e:
                        add_log(f"    ⚠️  Errore EMA20 series {isin}: {e}")

                    # CALCOLA SL SUGGERITO — formula ibrida
                    sl_data = analyzer.calculate_sl_suggerito_l1(entry_price, current_price, ema20)
                    sl_suggerito = sl_data.get('sl_suggerito')

                    # CALCOLA SG SUGGERITO — Stop Gain Dinamico basato su Slope (STEP 3 v3.0)
                    sg_data = analyzer.calculate_stop_gain_dynamic(entry_price, current_price, ema20_series, analyzer.p)
                    if sg_data.get('trigger'):
                        sg_suggerito = current_price  # Take profit immediato
                        add_log(f"    💰 STOP_GAIN_DYNAMIC triggered: {sg_data.get('reason')}")
                    else:
                        sg_suggerito = entry_price * (1 + sg_data.get('target_pct', 0.0))

                    # SEGNALI INFORMATIVI — SL/TP raggiunti (SOLO log/alert, MAI un
                    # automatismo di chiusura — stesso principio già applicato qui dal
                    # 2026-08-05, vedi fix gemello in _update_portfolio_l0_suggerito
                    # 2026-08-22 per il dettaglio completo). Il portafoglio reale non usa
                    # B/C/E/F/kill-switch (segnali "interni" della dashboard, non vendite):
                    # l'utente compra/vende manualmente su Directa, il monitor calcola solo
                    # cosa impostare — mai marca 'exited' da solo.
                    sl_hit = sl_suggerito is not None and current_price <= sl_suggerito
                    tp_hit = bool(sg_data.get('trigger'))

                    # Ratchet dello Stop tattico di avvicinamento al TP (order_pricing.py)
                    # — gira SEMPRE, anche oltre il TP (vedi commento gemello in
                    # _update_portfolio_l0_suggerito).
                    op = compute_order_prices(
                        current_price, sl_suggerito, sg_suggerito, broker,
                        previous_tightened_stop=float(prev_tp_stop_max) if prev_tp_stop_max else None,
                        sl_initial_pct=sl_initial_pct,
                        atr_pct=atr_pct,
                    )
                    tp_proximity_stop_max = op.get('tp_proximity_stop_max')
                    if op.get('tightened'):
                        add_log(f"    🔶 Stop tattico stretto a {op['prezzo_stop']:.2f}€ (avvicinamento/superamento TP)")
                        prev_val = float(prev_tp_stop_max) if prev_tp_stop_max else None
                        if prev_val is None or abs(tp_proximity_stop_max - prev_val) > 0.0001:
                            tightening_events.append({
                                'isin': isin, 'ticker': result.get('ticker'),
                                'fund_name': fund_name, 'variant': 'L1',
                                'current_price': current_price,
                                'prezzo_stop': op.get('prezzo_stop'),
                                'prezzo_limite_stop': op.get('prezzo_limite_stop'),
                                'tp_suggerito': sg_suggerito, 'broker': broker,
                            })

                    # Salva SL/SG suggerito — MAI 'exited' automaticamente
                    with conn.cursor() as cur:
                        cur.execute("""
                            UPDATE etf_portfolio_entries
                            SET sl_suggerito = %s, sg_suggerito = %s, stop_loss_updated_at = now(),
                                tp_proximity_stop_max = %s
                            WHERE id = %s
                        """, (sl_suggerito, sg_suggerito, tp_proximity_stop_max, entry_id))
                        conn.commit()

                    alert_msg = ''
                    if tp_hit:
                        alert_msg = ' 🟢 TARGET TP RAGGIUNTO — conferma uscita reale manualmente in dashboard'
                    elif sl_hit:
                        alert_msg = ' 🔴 SL RAGGIUNTO — conferma uscita reale manualmente in dashboard'

                    add_log(f"    {fund_name[:40]:40s} | SL: {sl_suggerito:.2f}€ | SG: {sg_suggerito:.2f}€{alert_msg}")

                except Exception as e:
                    add_log(f"    ⚠️  Errore L1 {isin}: {type(e).__name__}: {e}")
                    continue

            conn.close()

        except Exception as e:
            add_log(f"  ⚠️  Errore generale L1: {e}")

        return tightening_events

    def _update_portfolio_sl_suggested(self):
        """Aggiorna il stop loss suggerito per tutte le entry attive del portafoglio.
        Usa la formula continua basata su guadagno % e parametri della famiglia.
        """
        try:
            from technical_analysis import ETFTechnicalAnalyzer
            
            # Query: leggi tutte le entry attive
            query = "SELECT id, isin, entry_price, fund_name FROM etf_portfolio_entries WHERE status = 'active'"
            try:
                result = self.db.conn.execute(query)
                rows = result.fetchall()
            except:
                # Se la connessione è chiusa, non fare nulla
                return
            
            if not rows:
                return
            
            for entry_id, isin, entry_price, fund_name in rows:
                try:
                    entry_price = float(entry_price)
                    if entry_price <= 0:
                        continue
                    
                    # Leggi prezzo corrente
                    price_query = "SELECT close FROM etf_price_history WHERE isin = %s ORDER BY date DESC LIMIT 1"
                    price_result = self.db.conn.execute(price_query, (isin,))
                    price_row = price_result.fetchone()
                    
                    if not price_row:
                        continue
                    
                    current_price = float(price_row[0])
                    
                    # Calcola guadagno %
                    current_gain_pct = ((current_price - entry_price) / entry_price) * 100
                    
                    # Rileva famiglia da nome fondo
                    family = ETFTechnicalAnalyzer.detect_family(fund_name)
                    
                    # Crea analyzer e calcola SL
                    analyzer = ETFTechnicalAnalyzer(famiglia=family)
                    sl_result = analyzer.calculate_stop_loss(
                        current_price=current_price,
                        atr=None,
                        family_config=analyzer.p,
                        entry_price=entry_price,
                        current_gain_pct=current_gain_pct,
                        regime_str='BULL'
                    )
                    
                    # Prendi il SL trailing se disponibile, altrimenti initial
                    sl_suggested = sl_result.get('stop_loss_trailing') or sl_result.get('stop_loss_initial')
                    
                    if sl_suggested and sl_suggested > 0:
                        # Salva nel DB
                        update_query = "UPDATE etf_portfolio_entries SET stop_loss_suggested = %s, stop_loss_updated_at = now() WHERE id = %s"
                        self.db.conn.execute(update_query, (round(float(sl_suggested), 4), entry_id))
                        self.db.conn.commit()
                        
                except Exception as e:
                    add_log(f"  ⚠️  Errore SL per {isin}: {e}")
                    continue
                    
        except Exception as e:
            add_log(f"⚠️  Errore generale portfolio SL: {e}")


def main():
    monitor = ETFMonitor()
    monitor.run()


if __name__ == '__main__':
    main()

