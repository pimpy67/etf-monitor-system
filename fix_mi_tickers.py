#!/usr/bin/env python3
"""
fix_mi_tickers.py — Sostituisce i ticker .MI non funzionanti in etf_monitoraggio.xlsx
con ticker equivalenti su altre borse (preferibilmente .PA, .DE, .L, .AS).
"""
import openpyxl
import yfinance as yf
import time
import sys

EXCEL_PATH = "etf_monitoraggio.xlsx"

# Mappa COMPLETA: (row, old_ticker) → new_ticker
# Basata su lookup ISIN via Yahoo Finance + verifica yfinance
TICKER_MAP = {
    2:   ("KOR.MI",     "KRW.PA"),     # Amundi MSCI Korea (LU1900066975)
    3:   ("TUR.MI",     "TUR.PA"),     # Amundi MSCI Turkey (LU1900067601)
    4:   ("BRES.MI",    "LBRE.DE"),    # Amundi STOXX Europe 600 Basic Res Acc (LU1834983550)
    5:   ("GRC.MI",     "GRE.PA"),     # Amundi MSCI Greece (FR0010405431)
    6:   ("BRA.MI",     "RIO.PA"),     # Amundi MSCI Brazil Acc (LU1900066207)
    7:   ("BRES.MI",    "CHM.PA"),     # Amundi STOXX Europe 600 Basic Materials Acc (LU1834983634)
    8:   ("ENRG.MI",    "CWE.PA"),     # Amundi Global Bioenergy (LU1681046006)
    9:   ("ENER.MI",    "NRJ.PA"),     # Amundi MSCI New Energy Dist (FR0010524777)
    11:  ("CEC.MI",     "CEC.PA"),     # Amundi MSCI Eastern Europe Ex Russia (LU1900066462)
    12:  ("TELE.MI",    "TELE.PA"),    # Amundi STOXX Europe 600 Telecom (LU1834988609)
    13:  ("EMXC.MI",    "EMXC.L"),     # Amundi MSCI Emerging Ex China (LU2009202107)
    14:  ("UTIL.MI",    "UTI.MI"),     # Amundi STOXX Europe 600 Utilities — .MI funziona!
    15:  ("TPXH.MI",    "JPN.PA"),     # Amundi Japan TOPIX II Dis EUR Hdg (FR0010245514)
    16:  ("TPXH.MI",    "TPXH.PA"),   # Amundi IS Japan TOPIX Daily Hdg EUR (LU1681037864)
    17:  ("LAFRI.MI",   "LGQM.DE"),   # Amundi Pan Africa (LU1287022708)
    18:  ("AEEM.MI",    "AEMD.DE"),    # Amundi IS Core MSCI EM EUR Dist (LU1737652583)
    19:  ("AEEM.MI",    "AEME.PA"),    # Amundi IS Core MSCI EM Acc (LU1437017350)
    21:  ("TPXH.MI",    "TPXE.PA"),   # Amundi Japan TOPIX II Dis EUR (LU1681037609)
    22:  ("BNK.MI",     "BNK.PA"),     # Amundi STOXX Europe 600 Banks (LU1834983477)
    24:  ("AMEG.MI",    "AMEG.L"),     # Amundi MSCI EM ESG Broad Transition (LU2469335371)
    25:  ("CJ1.MI",     "CJ1.PA"),     # Amundi MSCI Japan ESG Broad Trans EUR Acc (LU1602144732)
    26:  ("AASI.MI",    "AEJ.PA"),     # Amundi MSCI AC Asia Pacific Ex Japan Acc (LU1900068328)
    27:  ("CRB.MI",     "COMO.PA"),    # Amundi Bl Equal-weight Comm ex-Agr (LU1829218749)
    28:  ("AASI.MI",    "EMAD.L"),     # Amundi IS MSCI EM Asia EUR (IE00B466KX20)
    30:  ("VLUE.MI",    "CV9.PA"),     # Amundi IS MSCI Europe Value Factor EUR (LU1681042518)
    31:  ("BNKE.MI",    "BNKE.PA"),    # Amundi Euro Stoxx Banks Acc (LU1829219390)
    33:  ("AASI.MI",    "APX.PA"),     # Amundi MSCI AC Asia Ex Japan Acc (LU1900068161)
    35:  ("WAT.MI",     "WAT.PA"),     # Amundi MSCI Water Dist (FR0010527275)
    36:  ("CI2.MI",     "LIGS.DE"),    # Amundi STOXX Europe 600 Industrials (LU1834987890)
    37:  ("FOOD.MI",    "LFOD.DE"),    # Amundi STOXX Europe 600 Consumer Staples (LU1834985845)
    38:  ("EMSR.MI",    "EMSRI.PA"),   # Amundi MSCI EM SRI Climate Acc (LU1861138961)
    39:  ("EMES.MI",    "SADM.DE"),    # Amundi MSCI EM ESG Selection Acc (LU2109787551)
    40:  ("EPRE.MI",    "EPRE.PA"),    # Amundi IS FTSE EPRA Europe Real Est EUR (LU1681039480)
    41:  ("CN1.MI",     "CN1.PA"),     # Amundi IS MSCI Nordic EUR (LU1681044647)
    42:  ("CEM.MI",     "MMS.PA"),     # Amundi MSCI EMU Small Cap ESG Dist (LU1598689153)
    43:  ("CV9.MI",     "VAL.PA"),     # Amundi MSCI EMU Value Factor Dist (LU1598690169)
    44:  ("CEU.MI",     "MEU.PA"),     # Amundi MSCI Europe Acc (FR0010261198)
    45:  ("CEU.MI",     "CEU2.PA"),    # Amundi IS Core MSCI Europe Acc (LU1437015735)
    46:  ("FMI.MI",     "MIB.PA"),     # Amundi FTSE MIB Acc (FR0010010827)
    47:  ("LWCE.MI",    "LWCE.PA"),    # Amundi IS MSCI Europe Climate Paris Acc (LU2130768844)
    48:  ("EUPA.MI",    "CEUG.DE"),    # Amundi MSCI Europe ESG Broad Trans Acc (LU1681042609)
    49:  ("RS2K.MI",    "RS2K.PA"),    # Amundi IS Russell 2000 EUR (LU1681038672)
    51:  ("EUPA.MI",    "ESGH.PA"),    # Amundi MSCI Europe ESG Sel EUR Hdg Acc (LU1940199984)
    52:  ("EUPA.MI",    "ESGE.PA"),    # Amundi MSCI Europe ESG Selection Acc (LU1940199711)
    53:  ("HLT.MI",     "HLT.PA"),     # Amundi STOXX Europe 600 Healthcare (LU1834986900)
    54:  ("MFDD.MI",    "MFDD.L"),     # Amundi IS MSCI EMU ESG Broad Trans Dist (LU0908501132)
    55:  ("QCEU.MI",    "QCEU.PA"),    # Amundi IS MSCI Europe Quality Factor EUR (LU1681041890)
    56:  ("CHIP.MI",    "CHIP.SW"),    # Amundi MSCI Semiconductors Acc (LU1900066033)
    57:  ("FMI.MI",     "FMI.PA"),     # Amundi IS Italy MIB ESG Acc (LU1681037518)
    58:  ("MSE.MI",     "MSE.PA"),     # Amundi EURO STOXX 50 II Acc (FR0007054358)
    60:  ("CMU.MI",     "CMU.PA"),     # Amundi MSCI EMU ESG Selection Acc (LU1602144575)
    61:  ("CD9.MI",     "DJE.PA"),     # Amundi DJ Industrial Average Dist (FR0007056841)
    62:  ("WINC.MI",    "WINC.AS"),    # iShares World Equity High Income Active (IE000KJPDY61)
    63:  ("CG9.MI",     "LGWT.DE"),    # Amundi IS MSCI Europe Growth Dist (LU1598688189)
    64:  ("EUPA.MI",    "EUSRI.PA"),   # Amundi MSCI Europe SRI Climate Acc (LU1861137484)
    65:  ("EPRA.MI",    "MWO.PA"),     # Amundi IS FTSE EPRA NAREIT Global II Dist (LU1832418773)
    66:  ("CC1.MI",     "CNAL.L"),     # Amundi MSCI China A (FR0011720911)
    67:  ("EUPA.MI",    "EDSRI.MI"),   # Amundi MSCI Europe SRI Paris Dist (LU2059756598) — keep .MI
    68:  ("ACWE.MI",    "ACWI.PA"),    # Amundi MSCI All Country World EUR Acc (LU1829220216)
    69:  ("EMI.MI",     "EMI.DE"),     # Amundi Euro Government IL Bond Acc (LU1650491282)
    71:  ("UINC.MI",    "INCI.AS"),    # iShares U.S. Equity High Income Active (IE0007FM00T9)
    72:  ("BTP10.MI",   "BTP10.MI"),   # BROKEN: no working alternative — keep for now
    73:  ("WOEE.MI",    "WOEE.AS"),    # iShares World Equity Enhanced Active USD (IE000D8XC064)
    74:  ("WHEA.MI",    "HLTW.PA"),    # Amundi MSCI World Health Care EUR Acc (LU0533033238)
    75:  ("X15E.MI",    "MTE.PA"),     # Amundi Euro Government Bond 10-15Y (LU1650489385)
    76:  ("IFLX.MI",    "IFLX.DE"),    # iShares EUR Flexible Income Bond Active (IE000NHAIBN0)
    77:  ("SWDA.MI",    "IWDA.L"),     # iShares Core MSCI World USD Acc (IE00B4L5Y983)
    78:  ("ECRP.MI",    "CC4.PA"),     # Amundi IS EUR Corporate Bond ESG 2 Acc (LU1681039647)
    80:  ("CW8.MI",     "WLDH.PA"),    # Amundi MSCI World Swap II EUR Hdg Dist (FR0011660927)
    81:  ("ECRP.MI",    "ECRP.PA"),    # Amundi IS EUR Corporate Bond ESG Acc (LU1437018168)
    82:  ("CC1.MI",     "CC1.PA"),     # Amundi MSCI China Tech EUR Acc (LU1681043912)
    83:  ("X710.MI",    "MTDD.DE"),    # Amundi Euro Government Bond 7-10Y (LU0290357259)
    84:  ("X1G.MI",     "X1G.PA"),     # Amundi IS Euro Lowest Rated IG Gov Bond (LU1681046774)
    85:  ("ECRP.MI",    "CRPX.L"),     # Amundi EUR Corp Bond Climate Paris Acc (LU1829219127)
    86:  ("EGRI.MI",    "EGRI.PA"),    # Amundi IS Euro Aggregate Bond ESG Acc (LU2182388236)
    87:  ("X57E.MI",    "X57E.DE"),    # Amundi Euro Government Bond 5-7Y (LU0290357176)
    88:  ("CW8.MI",     "CW8.PA"),     # Amundi IS MSCI World Swap EUR Acc (LU1681043599)
    89:  ("CW8.MI",     "WLD.PA"),     # Amundi MSCI World Swap II Dist (FR0010315770)
    90:  ("AHYE.MI",    "AHYE.PA"),    # Amundi IS Euro HY Bd ESG EUR (LU1681040496)
    91:  ("GOVA.MI",    "10AL.DE"),    # Amundi IS Core Euro Government Bond Dist (LU1737653714)
    92:  ("GGOV.MI",    "EAH.DE"),     # Amundi Euro Government Green Bond Acc (LU2356220926)
    93:  ("CW8.MI",     "WLDC.PA"),    # Amundi MSCI World Swap II Acc (FR0014003IY1)
    94:  ("CB3.MI",     "CB3.PA"),     # Amundi IS Euro Gov Tilted Green Bond Acc (LU1681046261)
    95:  ("GAGG.MI",    "CLIM.PA"),    # Amundi Global Aggr Green Bond Dist (LU1563454310)
    96:  ("CORP.MI",    "EBBB.PA"),    # Amundi IS EUR Corporate Bond 1-5Y ESG (LU1525418643)
    97:  ("CBEF.MI",    "CNB.PA"),     # Amundi EUR Corp Bond ex-Financials ESG (LU1829218822)
    98:  ("GHYA.MI",    "GHYU.L"),     # Amundi IS Global HY Corp Bond ESG Acc (LU2099295466)
    99:  ("ELCR.MI",    "ELCR.PA"),    # Amundi MSCI Smart Mobility Acc (LU2023679090)
    100: ("500H.MI",    "S500.PA"),    # Amundi S&P 500 Screened EUR Hdg Acc (IE000KXCEXR3)
    101: ("GOVA.MI",    "EGOV.PA"),    # Amundi IS Core Euro Government Bond Acc (LU1437018598)
    102: ("X35E.MI",    "EGV5.DE"),    # Amundi Euro Government Bond 3-5Y (LU1650488494)
    103: ("WOEH.MI",    "WOEH.MI"),    # iShares World Equity Enhanced EUR Hdg — .MI funziona!
    104: ("X15E.MI",    "E15G.DE"),    # Amundi Euro Government Bond 15+Y (LU1287023268)
    105: ("BTP13.MI",   "BTP13.MI"),   # BROKEN: no working alternative — keep for now
    106: ("MA35.MI",    "MA35.PA"),    # Amundi Euro Highest Rated MW Gov Bond 3-5Y (LU1829219713)
    107: ("AM3A.MI",    "AM3A.PA"),    # Amundi IS Euro Gov Bond Highest Rated IG Acc (LU1681046691)
    109: ("EFRN.MI",    "EFRN.DE"),    # Amundi IS EUR Float Rate Corp Bd ESG (IE00BF5GB717)
    110: ("USYH.MI",    "LYXE.DE"),    # Amundi IS USD HY Corp Bond ESG EUR Hdg Dist (LU1435356495)
    111: ("X1G.MI",     "X13G.PA"),    # Amundi IS Euro Lowest Rated IG Gov Bond 1-3Y (LU1681046345)
    112: ("MWRD.MI",    "MGT.PA"),     # Amundi DJ Global Titans 50 Dist (FR0007075494)
    113: ("X13E.MI",    "MTA.PA"),     # Amundi Euro Government Bond 1-3Y (LU1650487413)
    114: ("MA13.MI",    "MA13.PA"),    # Amundi Euro Highest Rated MW Gov Bond 1-3Y (LU1829219556)
    115: ("STHY.MI",    "YIEL.L"),     # Amundi IS EUR ST HY Corporate Bond ESG Dist (LU1812090543)
    116: ("ECRP.MI",    "ECR3.DE"),    # Amundi IS EUR Corporate Bond 0-3Y ESG Acc (LU2037748774)
    117: ("USYH.MI",    "USHY.L"),     # Amundi IS USD HY Corp Bond ESG Dist (LU1435356149)
    118: ("FINSW.MI",   "FINSW.PA"),   # Amundi MSCI World Financials EUR Acc (LU0533032859)
    119: ("C40.MI",     "CACC.PA"),    # Amundi IS CAC 40 ESG Acc (FR0013380607)
    120: ("LWCR.MI",    "LWCR.DE"),    # Amundi MSCI World ESG Broad Transition Acc (IE0001GSQ2O9)
    121: ("AFLT.MI",    "AFLT.PA"),    # Amundi IS USD Float Rate Corp Bd ESG (LU1681040900)
    122: ("ERNA.MI",    "YCSH.DE"),    # iShares EUR Cash UCITS ETF EUR Acc (IE000JJPY166)
    123: ("C3M.MI",     "C3M.PA"),     # Amundi Euro Government Bond 0-6M Acc (FR0010754200)
    124: ("CSH.MI",     "CSH.PA"),     # Amundi EUR Overnight Return (FR0010510800)
    126: ("AGGH.MI",    "GAGG.PA"),    # Amundi IS Core Global Aggregate Bond Acc (LU1437024729)
    127: ("CORP.MI",    "UBBB.PA"),    # Amundi IS Global Corporate Bond 1-5Y ESG Acc (LU1525418726)
    129: ("EPRE.MI",    "EPRE.L"),     # Amundi IS FTSE EPRA Europe Real Est Dist (LU1812091194)
    130: ("CRB.MI",     "COMH.MI"),    # Amundi Bl Equal-weight Comm ex-Agr EUR H — .MI funziona!
    131: ("USYH.MI",    "UHYC.L"),     # Amundi IS USD HY Corp Bond ESG Acc (LU1435356065)
    132: ("EMBHI.MI",   "EMBHI.MI"),   # Amundi IS USD EM Gov Bond EUR Hdg Dist — .MI funziona!
    133: ("AGEB.MI",    "AGEB.PA"),    # Amundi IS USD EM Gov Bond Acc (LU1681041205)
    134: ("USEH.MI",    "USEH.MI"),    # iShares US Equity Enhanced EUR Hdg — .MI funziona!
    136: ("500H.MI",    "500H.PA"),    # Amundi IS S&P 500 Swap EUR Hdg Acc (LU1681049109)
    137: ("LEMB.MI",    "LEMB.L"),     # Amundi IS USD EM Gov Bond Dist (LU1686830909)
    138: ("500H.MI",    "500.PA"),     # Amundi IS S&P 500 Swap EUR Acc (LU1681048804)
    139: ("USIH.MI",    "USIH.PA"),    # Amundi USD Corp Bond Climate Paris EUR Hdg Dist (LU1285960032)
    140: ("PRAP.MI",    "PRAP.SW"),    # Amundi IS USD Corporate Bond ESG Acc (LU2089239276)
    141: ("GGOV.MI",    "GOVH.PA"),    # Amundi IS Core Global Gov Bond EUR Hdg Acc (LU1708330235)
    142: ("USIG.MI",    "USIG.L"),     # Amundi USD Corp Bond Climate Paris Dist (LU1285959703)
    143: ("GGOV.MI",    "GGOV.PA"),    # Amundi IS Core Global Gov Bond Acc (LU1437016204)
    144: ("USIC.MI",    "USIC.L"),     # Amundi USD Corp Bond Climate Paris Acc (LU1285959885)
    145: ("USRIH.MI",   "USRIH.MI"),   # BROKEN: no alternative found — keep for now
    147: ("TNO.MI",     "TNO.MI"),     # Amundi STOXX Europe 600 Technology — .MI funziona!
    148: ("CC1.MI",     "LHKG.DE"),    # Amundi MSCI China ESG Selection Extra Dist (LU1900067940)
    149: ("CC1.MI",     "ASI.PA"),     # Amundi MSCI China ESG Selection Extra Acc (LU1900068914)
    151: ("USEE.MI",    "USEE.AS"),    # iShares US Equity Enhanced Active USD (IE0009VWHAE6)
    152: ("SADU.MI",    "SADU.PA"),    # Amundi MSCI USA ESG Selection Acc (IE000PEAJOT0)
    153: ("CATH.MI",    "WLSC.PA"),    # Amundi MSCI World Catholic Principles Acc (IE000QWCYQT0)
    154: ("INDI.MI",    "INR.PA"),     # Amundi MSCI India Swap EUR Acc (FR0010361683)
    155: ("INS.MI",     "LIRU.DE"),    # Amundi STOXX Europe 600 Insurance (LU1834987973)
    156: ("INDI.MI",    "CI2.PA"),     # Amundi IS MSCI India Swap II EUR Acc (LU1681043086)
    157: ("GLUX.MI",    "GLUX.PA"),    # Amundi IS Global Luxury EUR Acc (LU1681048630)
    158: ("UST.MI",     "NDXH.PA"),    # Amundi IS Nasdaq-100 Swap EUR Hdg Acc (LU1681038599)
    159: ("UST.MI",     "UST.PA"),     # Amundi Core Nasdaq-100 Swap EUR Hdg Acc (LU1954152853)
    160: ("DTEC.MI",    "UNIC.L"),     # Amundi MSCI Disruptive Technology Acc (LU2023678282)
    161: ("UST.MI",     "UST.PA"),     # Amundi Core Nasdaq-100 Swap Acc (LU1829221024)
    162: ("UST.MI",     "ANX.PA"),     # Amundi IS Nasdaq-100 Swap EUR Acc (LU1681038243)
    163: ("FOOD.MI",    "LTVL.DE"),    # Amundi STOXX Europe 600 Consumer Discr (LU1834988781)
    164: ("GENY.MI",    "GENY.L"),     # Amundi MSCI Millennials Acc (LU2023678449)
    165: ("GOAI.MI",    "GOAI.PA"),    # Amundi MSCI Robotics & AI Acc (LU1861132840)
    166: ("INDO.MI",    "INDO.PA"),    # Amundi MSCI Indonesia Acc (LU1900065811)
    167: ("WTEC.MI",    "TNOW.PA"),    # Amundi MSCI World Information Tech EUR Acc (LU0533033667)
    168: ("DIGE.MI",    "DIGE.L"),     # Amundi MSCI Digital Economy Acc (LU2023678878)
    169: ("DFNS.MI",    "DEFS.PA"),    # Amundi IS Stoxx Europe Defense Acc (LU3038520774)
    170: ("PHAU.MI",    "PHAU.L"),     # WisdomTree Physical Gold ETC (JE00B1VS3770)
    171: ("PHAG.MI",    "PHAG.L"),     # WisdomTree Physical Silver ETC (JE00B1VS3333)
    172: ("VWCE.MI",    "VWRA.L"),     # Vanguard FTSE All-World USD Acc (IE00BK5BQT80)
    173: ("IWDP.MI",    "IASP.SW"),    # iShares Developed Markets Property Yield (IE00B1FZS244)
    174: ("IPRP.MI",    "IWDP.SW"),    # iShares European Property Yield EUR Dist (IE00B1FZS350)
    175: ("INRG.MI",    "INRG.SW"),    # iShares Global Clean Energy USD Dist (IE00B1XNHC34)
    177: ("CSCA.MI",    "SXR2.DE"),    # iShares MSCI Canada USD Acc (IE00B52SF786)
    178: ("INFR.MI",    "IQQI.DE"),    # iShares Global Infrastructure USD Dist (IE00B6QW5T61)
    179: ("IBTM.MI",    "IBGX.AS"),    # iShares USD Treasury Bond 7-10yr USD Dist (IE00B1FZS681)
    182: ("VHYL.MI",    "VHYL.L"),     # Vanguard FTSE All-World High Div Yield
    187: ("LTAM.MI",    "LTAM.L"),     # iShares MSCI EM Latin America USD Dist
    188: ("EMVL.MI",    "EMVL.L"),     # iShares MSCI EM Min Vol Factor USD Acc
    190: ("EMID.MI",    "EMID.L"),     # iShares MSCI Europe Mid Cap EUR Dist
    191: ("AIGI.MI",    "AIGI.L"),     # WisdomTree Industrial Metals ETC (GB00B15KYG56)
    192: ("COPA.MI",    "COPA.L"),     # WisdomTree Copper ETC (GB00B15KXQ89)
    193: ("ALUM.MI",    "ALUM.L"),     # WisdomTree Aluminium ETC
    194: ("ZINC.MI",    "ZINC.L"),     # WisdomTree Zinc ETC
    195: ("MINE.MI",    "COPM.AS"),    # iShares Copper Miners USD Acc (IE00063FT9K6)
    196: ("BATG.MI",    "BATG.L"),     # L&G Battery Value-Chain
    197: ("IPRV.MI",    "IDPE.L"),     # iShares Listed Private Equity USD Dist (IE00B1TXHL60)
    198: ("XLPE.MI",    "XLPE.L"),     # Xtrackers LPX Private Equity Swap 1C
    199: ("XEON.MI",    "XEON.DE"),    # Xtrackers II EUR Overnight Rate Swap (LU0290358497)
    200: ("IEGE.MI",    "DLTM.L"),     # iShares EUR Govt Bond 0-1yr EUR Acc (IE00B27YCK28)
    201: ("IUSE.MI",    "IUSE.L"),     # iShares Core S&P 500 EUR Hedged Acc (IE00B3ZW0K18)
    202: ("IWDE.MI",    "IWDE.L"),     # iShares MSCI World EUR Hedged Acc (IE00B441G979)
}

def verify_ticker(ticker, fast=True):
    """Verifica che il ticker abbia dati recenti su Yahoo Finance."""
    try:
        t = yf.Ticker(ticker)
        hist = t.history(period='10d')
        if len(hist) == 0:
            return False, "no history"
        prices = hist['Close'].dropna()
        if len(prices) == 0:
            return False, "all nan"
        last_date = str(hist.index[-1].date())
        last_price = prices.iloc[-1]
        return True, f"{last_date} @ {last_price:.2f}"
    except Exception as e:
        return False, str(e)

def apply_fix(verify=True):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['ETF']

    changed = []
    failed_verify = []
    kept_broken = []

    for row_num, (old_ticker, new_ticker) in TICKER_MAP.items():
        current = ws.cell(row_num, 2).value
        nome = ws.cell(row_num, 3).value or ''

        if current == new_ticker:
            print(f"Row {row_num}: {current} → già corretto")
            continue

        if old_ticker != new_ticker:
            # Nuova mappa: cambia ticker
            if verify:
                ok, info = verify_ticker(new_ticker)
                if not ok:
                    # Ticker nuovo non funziona — mantieni il vecchio .MI e segnala
                    failed_verify.append((row_num, old_ticker, new_ticker, info, nome[:50]))
                    print(f"Row {row_num}: {old_ticker} → {new_ticker} FALLITO ({info}) — mantenuto .MI")
                    continue
                print(f"Row {row_num}: {old_ticker} → {new_ticker} OK [{info}]")
            else:
                print(f"Row {row_num}: {old_ticker} → {new_ticker} (no verify)")

            ws.cell(row_num, 2).value = new_ticker
            changed.append((row_num, old_ticker, new_ticker, nome[:50]))
        else:
            kept_broken.append((row_num, old_ticker, nome[:50]))

        time.sleep(0.3)

    wb.save(EXCEL_PATH)

    print(f"\n=== RIEPILOGO ===")
    print(f"Ticker aggiornati: {len(changed)}")
    print(f"Ticker falliti verifica: {len(failed_verify)}")
    print(f"Ticker mantenuti (senza alternativa): {len(kept_broken)}")

    if failed_verify:
        print("\nFALLITI (da rivedere):")
        for row, old, new, info, nome in failed_verify:
            print(f"  Row {row}: {old} → {new}: {info} [{nome}]")

    if kept_broken:
        print("\nMANTENUTI (nessuna alternativa trovata):")
        for row, t, nome in kept_broken:
            print(f"  Row {row}: {t} [{nome}]")

    return changed, failed_verify, kept_broken

if __name__ == '__main__':
    verify = '--no-verify' not in sys.argv
    print(f"Avvio fix ticker .MI → alternative (verify={verify})...")
    print(f"Excel: {EXCEL_PATH}")
    print()
    apply_fix(verify=verify)
