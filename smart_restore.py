#!/usr/bin/env python3
"""
smart_restore.py — Ripristina selettivamente i dati "live" dall'Excel di backup VPS
nel nuovo Excel proveniente da git, preservando i ticker e le configurazioni aggiornate.

Colonne ripristinate dal backup:
  - Colonna 1: Livello (L0/L1/L2/L3 — stato corrente del monitoring)

Tutto il resto (Ticker, Nome, Categoria, Borsa, Valuta, Prezzo, indicatori, ISIN)
viene dal git — questo permette alle correzioni dei ticker di passare.

Uso: python3 smart_restore.py <backup_excel> <git_excel> [output_excel]
Se output_excel non specificato, sovrascrive git_excel.
"""
import sys
import openpyxl

def smart_restore(backup_path, git_path, output_path=None):
    if output_path is None:
        output_path = git_path

    try:
        wb_backup = openpyxl.load_workbook(backup_path)
        ws_backup = wb_backup['ETF']
    except Exception as e:
        print(f"ATTENZIONE: impossibile aprire backup ({e}) — uso Excel git invariato")
        return

    try:
        wb_git = openpyxl.load_workbook(git_path)
        ws_git = wb_git['ETF']
    except Exception as e:
        print(f"ERRORE: impossibile aprire Excel git: {e}")
        sys.exit(1)

    # Mappa ISIN (colonna 15) → Livello (colonna 1) dal backup VPS
    # Se ISIN è vuoto, usa Ticker (colonna 2) come chiave fallback
    backup_levels = {}
    for r in range(2, ws_backup.max_row + 1):
        lvl = ws_backup.cell(r, 1).value
        isin = ws_backup.cell(r, 15).value
        ticker = ws_backup.cell(r, 2).value
        key = str(isin).strip() if isin and str(isin).strip() not in ('', 'None') else f'TICKER:{ticker}'
        if key and lvl is not None:
            backup_levels[key] = lvl

    restored = 0
    skipped = 0
    for r in range(2, ws_git.max_row + 1):
        isin = ws_git.cell(r, 15).value
        ticker = ws_git.cell(r, 2).value
        key = str(isin).strip() if isin and str(isin).strip() not in ('', 'None') else f'TICKER:{ticker}'

        if key in backup_levels:
            ws_git.cell(r, 1).value = backup_levels[key]
            restored += 1
        else:
            skipped += 1

    wb_git.save(output_path)
    print(f"smart_restore: {restored} Livelli ripristinati, {skipped} nuovi ETF (senza backup) mantenuti invariati")
    print(f"Output: {output_path}")

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print(f"Uso: {sys.argv[0]} <backup_excel> <git_excel> [output_excel]")
        sys.exit(1)
    backup = sys.argv[1]
    git_excel = sys.argv[2]
    output = sys.argv[3] if len(sys.argv) > 3 else None
    smart_restore(backup, git_excel, output)
